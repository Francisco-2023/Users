from flask import Flask, render_template, redirect, url_for, flash, current_app, session
import subprocess, socket, requests, dns.resolver, whois, speedtest
from routeros_api import RouterOsApiPool
from flask import request, jsonify
import platform
from functools import wraps
import subprocess
import string
from werkzeug.security import generate_password_hash
import re, psycopg2, requests
import os 
from werkzeug.utils import secure_filename
from datetime import date
from datetime import datetime
from psycopg2.extras import RealDictCursor
import calendar
import json
import speedtest
app = Flask(__name__)
app.secret_key = 'tu_clave_secreta_aqui_que_debe_ser_larga_y_segura'


# ✅ Define la carpeta donde se guardarán las imágenes
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ✅ Crea la carpeta si no existe
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

fecha_inicio = date.today()
hoy = date.today()
if hoy.day <= 5:
    fecha_corte = date(hoy.year, hoy.month, 5)
else:
    # Si ya pasó el 5, la fecha de corte será el 5 del siguiente mes
    mes = hoy.month + 1 if hoy.month < 12 else 1
    año = hoy.year if mes != 1 else hoy.year + 1
    fecha_corte = date(año, mes, 5)

def verificar_suspension(cliente):
    hoy = date.today()
    if not cliente.pagado and hoy > cliente.fecha_corte:
        cliente.fecha_suspension = hoy

# 📌 CONFIGURACIÓN DEL BOT TELEGRAM
TELEGRAM_BOT_TOKEN = "7642549213:AAEmh55fiDaMzKWPpd52BDuQIGSSaqoMBwA"
TELEGRAM_CHAT_ID = "6193243594"



def enviar_mensaje_telegram(mensaje):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    data = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": mensaje,
        "parse_mode": "Markdown"
    }

    try:
        response = requests.post(url, data=data)
        if response.status_code == 200:
            print("✅ Mensaje enviado correctamente a Telegram.")
        else:
            print(f" Error al enviar mensaje a Telegram. Código: {response.status_code}")
            print("Respuesta:", response.text)
    except Exception as e:
        print(" Excepción al enviar mensaje a Telegram:", e)

@app.route('/obtener_chat_id', methods=['GET'])
def obtener_chat_id():
    updates = requests.get(f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates').json()
    return updates


# 📍 VER PÁGINA DE TICKETS
@app.route('/ticket')
def ticket():
    conn = get_db_connection()
    cur = conn.cursor()

    # JOIN con reporte_tecnico para obtener el id del reporte si existe
    cur.execute("""
        SELECT t.*, r.id AS reporte_id
        FROM tickets t
        LEFT JOIN reporte_tecnico r ON r.ticket_id = t.id
        ORDER BY t.fecha DESC
    """)
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    tickets = [dict(zip(columnas, fila)) for fila in filas]

    # Obtener contratos activos
    cur.execute("SELECT numero_contrato, nombres, apellidos FROM contratos WHERE estado = 'Activo'")
    contratos_filas = cur.fetchall()
    columnas_contratos = [desc[0] for desc in cur.description]
    contratos = [dict(zip(columnas_contratos, fila)) for fila in contratos_filas]

    cur.close()
    conn.close()

    return render_template("ticket.html", tickets=tickets, contratos=contratos)


# 🧾 ENDPOINT PARA CREAR TICKET
@app.route('/crear_ticket', methods=['POST'])
def crear_ticket():
    usuario = session.get('username', 'Anónimo')  # o 'Desconocido' si quieres un valor por defecto
    numero_contrato = request.form['numero_contrato']
    phone = request.form.get('phone')
    address_reference = request.form.get('address_reference')
    descripcion = request.form['descripcion']
    categoria = request.form.get('categoria')
    prioridad = request.form.get('prioridad')
    assigned_to = request.form.get('assigned_to')

    # Insertar en base de datos
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO tickets (usuario, numero_contrato, phone, address_reference, descripcion, categoria, prioridad, assigned_to)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (usuario, numero_contrato, phone, address_reference, descripcion, categoria, prioridad, assigned_to))

    conn.commit()
    cur.close()
    conn.close()

    # Enviar mensaje a Telegram
    mensaje = f"""🎫 *Nuevo Ticket Creado* 🎫

👤 *Usuario:* {usuario}
🔗 *Contrato:* {numero_contrato}
📞 *Teléfono:* {phone or 'N/A'}
📍 *Referencia:* {address_reference or 'N/A'}
📝 *Descripción:* {descripcion}
📂 *Categoría:* {categoria or 'No especificada'}
⚠️ *Prioridad:* {prioridad or 'Normal'}
👷 *Asignado a:* {assigned_to or 'Pendiente'}
"""
    enviar_mensaje_telegram(mensaje)

    flash('Ticket creado exitosamente y enviado a Telegram ✅', 'success')
    return redirect(url_for('ticket'))


# 🛠️ ACTUALIZAR TICKET
@app.route('/actualizar_ticket/<int:ticket_id>', methods=['POST'])
def actualizar_ticket(ticket_id):
    usuario = request.form['usuario']
    numero_contrato = request.form['numero_contrato']
    phone = request.form.get('phone')
    address_reference = request.form.get('address_reference')
    descripcion = request.form['descripcion']
    categoria = request.form.get('categoria')
    prioridad = request.form.get('prioridad')
    assigned_to = request.form.get('assigned_to')
    estado = request.form.get('estado')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE tickets
        SET usuario = %s,
            numero_contrato = %s,
            phone = %s,
            address_reference = %s,
            descripcion = %s,
            categoria = %s,
            prioridad = %s,
            assigned_to = %s,
            estado = %s
        WHERE id = %s
    """, (usuario, numero_contrato, phone, address_reference, descripcion, categoria, prioridad, assigned_to, estado, ticket_id))

    conn.commit()
    cur.close()
    conn.close()

    flash('Ticket actualizado correctamente ✅', 'success')
    return redirect(url_for('ticket'))


# 🗑️ ELIMINAR TICKET
@app.route('/eliminar_ticket/<int:ticket_id>', methods=['POST'])
def eliminar_ticket(ticket_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM tickets WHERE id = %s", (ticket_id,))
    conn.commit()
    cur.close()
    conn.close()

    flash('Ticket eliminado correctamente 🗑️', 'warning')
    return redirect(url_for('ticket'))


# 🟢 ACTUALIZAR ESTADO (Solo estado)
@app.route('/actualizar_estado/<int:ticket_id>', methods=['POST'])
def actualizar_estado(ticket_id):
    estado = request.form.get('estado')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE tickets SET estado = %s WHERE id = %s", (estado, ticket_id))
    conn.commit()
    cur.close()
    conn.close()

    flash('Estado del ticket actualizado ✅', 'success')
    return redirect(url_for('ticket'))


@app.route('/reporte/<int:ticket_id>', methods=['GET', 'POST'])
def reporte(ticket_id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        # ✅ Obtener datos del formulario
        descripcion = request.form['descripcion']
        materiales = request.form.getlist('materiales[]')
        cantidades = request.form.getlist('cantidades[]')
        equipo_usado_id = request.form.get('equipo_usado')
        latitud = request.form.get('latitud')
        longitud = request.form.get('longitud')
        foto = request.files.get('foto')
        nombre_foto = ''
        fecha = datetime.now()

        # Obtener usuario logueado de la sesión
        usuario = session.get('username', 'desconocido')

        # ✅ Validar foto y guardar
        if foto and allowed_file(foto.filename):
            nombre_foto = secure_filename(f"{fecha.strftime('%Y%m%d%H%M%S')}_{foto.filename}")
            ruta_foto = os.path.join(app.config['UPLOAD_FOLDER'], nombre_foto)
            foto.save(ruta_foto)

        # ✅ Unir materiales con cantidades
        materiales_usados = [
            f"{mat} ({cant})" for mat, cant in zip(materiales, cantidades)
        ]
        materiales_texto = ", ".join(materiales_usados)

        # ✅ Validar relación con detalles_producto
        cur.execute("SELECT id FROM detalles_producto WHERE id = %s", (equipo_usado_id,))
        equipo_valido = cur.fetchone()
        if not equipo_valido:
            flash("⚠️ El equipo seleccionado no existe en los detalles del inventario.", "danger")
            cur.close()
            conn.close()
            return redirect(url_for('reporte', ticket_id=ticket_id))

        # ✅ Insertar reporte con usuario
        cur.execute("""
            INSERT INTO reporte_tecnico (
                ticket_id, descripcion, materiales, equipo_usado_id, latitud, longitud, foto, fecha, usuario
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            ticket_id, descripcion, materiales_texto, equipo_usado_id,
            latitud, longitud, nombre_foto, fecha, usuario
        ))

        # ✅ Actualizar estado del ticket a 'Hecho'
        cur.execute("""
            UPDATE tickets
            SET estado = 'Hecho'
            WHERE id = %s
        """, (ticket_id,))

        conn.commit()
        flash("✅ Reporte guardado correctamente", "success")
        cur.close()
        conn.close()
        return redirect(url_for('ticket'))

    # ✅ Obtener ticket
    cur.execute("SELECT * FROM tickets WHERE id = %s", (ticket_id,))
    ticket = cur.fetchone()

    # ✅ Obtener inventario con detalles del producto
    cur.execute("""
        SELECT i.id, i.nombre, i.marca, i.modelo, i.codigo, d.id, d.numero_serie, d.mac_address
        FROM inventario i
        LEFT JOIN detalles_producto d ON i.id = d.producto_id
        WHERE i.estado IN ('Disponible', 'En uso')
        ORDER BY i.nombre
    """)
    filas = cur.fetchall()

    # ✅ Agrupar inventario y detalles
    inventario_agrupado = {}
    for fila in filas:
        pid = fila[0]  # inventario.id
        if pid not in inventario_agrupado:
            inventario_agrupado[pid] = {
                'id': pid,
                'nombre': fila[1],
                'marca': fila[2],
                'modelo': fila[3],
                'codigo': fila[4],
                'detalles': []
            }
        if fila[5]:  # detalle_id no es None
            inventario_agrupado[pid]['detalles'].append({
                'id': fila[5],
                'serie': fila[6],
                'mac': fila[7]
            })

    cur.close()
    conn.close()

    return render_template(
        "reporte.html",
        ticket=ticket,
        inventario=list(inventario_agrupado.values()),
        datetime=datetime
    )


@app.route('/reporte/ver/<int:reporte_id>')
def ver_reporte(reporte_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT r.id, r.ticket_id, r.descripcion, r.materiales, r.latitud, r.longitud, r.foto, r.fecha, r.usuario,
               d.numero_serie, d.mac_address, i.nombre AS equipo_nombre, i.marca, i.modelo 
        FROM reporte_tecnico r
        LEFT JOIN detalles_producto d ON r.equipo_usado_id = d.id
        LEFT JOIN inventario i ON d.producto_id = i.id
        WHERE r.id = %s
    """, (reporte_id,))
    reporte = cur.fetchone()
    cur.close()
    conn.close()

    if not reporte:
        flash("⚠️ El reporte técnico no existe.", "warning")
        return redirect(url_for('ticket'))

    return render_template("ver_reporte.html", reporte=reporte)



def cortar_contrato_mikrotik(ip):
    try:
        pool, api = conectar_api()
        queues = api.get_resource('/queue/simple')

        # Buscar el queue con la IP asignada y deshabilitarlo o eliminarlo para cortar
        queue_items = queues.get()
        for item in queue_items:
            if item.get('target') == f"{ip}/32":
                # Opción 1: Deshabilitar queue
                queues.update(id=item['.id'], disabled='true')
                # Opción 2: Eliminar queue (si prefieres)
                # queues.remove(id=item['.id'])
                pool.disconnect()
                return True
        pool.disconnect()
        print("No se encontró queue para la IP:", ip)
        return False
    except Exception as e:
        print("Error al conectar con MikroTik:", e)
        return False

def verificar_y_cortar_contrato(contrato_id, conn):
    cur = conn.cursor()
    hoy = date.today()

    # Obtener datos del contrato
    cur.execute("SELECT pagado, fecha_corte, estado, ip_asignada FROM contratos WHERE id = %s", (contrato_id,))
    contrato = cur.fetchone()

    if not contrato:
        return False  # Contrato no encontrado

    pagado, fecha_corte, estado, ip = contrato

    if not pagado and fecha_corte and hoy > fecha_corte and estado != 'Cortado':
        exito = cortar_contrato_mikrotik(ip)
        if exito:
            cur.execute("UPDATE contratos SET estado = %s, fecha_suspension = %s WHERE id = %s", ('Cortado', hoy, contrato_id))
            conn.commit()
            return True
        else:
            print("No se pudo cortar el contrato en MikroTik")
    return False

# Datos MikroTik
MIKROTIK_IP = '192.168.88.1'
USUARIO = 'Francisco'
CONTRASENA = '1251301881'

def conectar_api():
    pool = RouterOsApiPool(MIKROTIK_IP, username=USUARIO, password=CONTRASENA, plaintext_login=True)
    return pool, pool.get_api()

# Configuración base de datos
DB_CONFIG = {
    'host': 'localhost',
    'database': 'mikrotik_db',
    'user': 'postgres',
    'password': '0984591216'
}

def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)

def obtener_todas_las_ips_mikrotik():
    try:
        connection = RouterOsApiPool(
            host='192.168.88.1',  # Cambia por tu IP
            username='Francisco',
            password='1251301881',
            port=8728,
            plaintext_login=True
        )
        api = connection.get_api()

        # IPs de clientes DHCP
        dhcp_clients = api.get_resource('/ip/dhcp-server/lease').get()
        ips_dhcp = [client.get('address') for client in dhcp_clients if 'address' in client]

        # IPs en la tabla ARP
        arp_entries = api.get_resource('/ip/arp').get()
        ips_arp = [entry.get('address') for entry in arp_entries if 'address' in entry]

        # IPs de vecinos (mndp/llpd)
        neighbors = api.get_resource('/ip/neighbor').get()
        ips_neighbors = [n.get('address') for n in neighbors if 'address' in n]

        # Combinamos y quitamos duplicados
        todas_ips = list(set(ips_dhcp + ips_arp + ips_neighbors))
        connection.disconnect()

        return sorted([ip for ip in todas_ips if ip is not None])
    except Exception as e:
        print(f"Error al obtener IPs: {e}")
        return []


# Supongamos que tienes el diccionario provincias definido en Python como en JS
provincias_list = [
    "AZUAY", "BOLÍVAR", "CAÑAR", "CARCHI", "COTOPAXI", "CHIMBORAZO", "EL ORO",
    "ESMERALDAS", "GUAYAS", "IMBABURA", "LOJA", "LOS RÍOS", "MANABÍ", "MORONA SANTIAGO",
    "NAPO", "PASTAZA", "PICHINCHA", "TUNGURAHUA", "ZAMORA CHINCHIPE", "GALÁPAGOS",
    "SUCUMBÍOS", "ORELLANA", "SANTO DOMINGO DE LOS TSÁCHILAS", "SANTA ELENA", "ZONA NO DELIMITADA"
]

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
@login_required
def index():
    # Verificar si hay sesión activa
    if 'username' not in session:
        return redirect(url_for('login'))

    pool, api = conectar_api()

    interfaces = api.get_resource('/interface').get()
    clientes_pppoe = api.get_resource('/ppp/active').get()
    clientes_hotspot = api.get_resource('/ip/hotspot/active').get()
    reglas_firewall = api.get_resource('/ip/firewall/filter').get()
    logs = api.get_resource('/log').get()
    logs = logs[-10:]

    interfaces_active = sum(1 for i in interfaces if i.get('running'))
    interfaces_total = len(interfaces)
    clientes_conectados = len(clientes_pppoe) + len(clientes_hotspot)
    reglas_activas = sum(1 for r in reglas_firewall if not r.get('disabled', False))
    reglas_desactivadas = sum(1 for r in reglas_firewall if r.get('disabled', False))

    monitor = api.get_resource('/interface/monitor-traffic')
    trafico_rx_bps_total = 0
    trafico_tx_bps_total = 0
    for iface in interfaces:
        nombre = iface.get('name')
        if nombre:
            try:
                datos = monitor.call('monitor', {'interface': nombre, 'once': True})
                if datos and len(datos) > 0:
                    trafico_rx_bps_total += int(datos[0].get('rx-bits-per-second', 0))
                    trafico_tx_bps_total += int(datos[0].get('tx-bits-per-second', 0))
            except:
                pass

    def formatear_bits(bps):
        if bps > 1e9:
            return f"{bps/1e9:.2f} Gbps"
        elif bps > 1e6:
            return f"{bps/1e6:.2f} Mbps"
        elif bps > 1e3:
            return f"{bps/1e3:.2f} Kbps"
        else:
            return f"{bps} bps"

    trafico_subida = formatear_bits(trafico_tx_bps_total)
    trafico_bajada = formatear_bits(trafico_rx_bps_total)

    pool.disconnect()

    return render_template('index.html',
                           interfaces_active=interfaces_active,
                           interfaces_total=interfaces_total,
                           clientes_conectados=clientes_conectados,
                           clientes_pppoe=len(clientes_pppoe),
                           clientes_hotspot=len(clientes_hotspot),
                           reglas_activas=reglas_activas,
                           reglas_desactivadas=reglas_desactivadas,
                           trafico_subida=trafico_subida,
                           trafico_bajada=trafico_bajada,
                           logs_recientes=logs)

# Endpoint: tráfico interfaces en tiempo real
@app.route('/api/trafico_todas')
def api_trafico_todas():
    pool, api = conectar_api()
    monitor = api.get_resource('/interface/monitor-traffic')
    interfaces = api.get_resource('/interface').get()

    resultados = {}
    for iface in interfaces:
        name = iface.get('name')
        if not name:
            continue
        try:
            data = monitor.call('monitor', {'interface': name, 'once': True})
            if data and len(data) > 0:
                rx_bps = int(data[0].get('rx-bits-per-second', 0))
                tx_bps = int(data[0].get('tx-bits-per-second', 0))
                resultados[name] = {'rx_bps': rx_bps, 'tx_bps': tx_bps}
        except:
            resultados[name] = {'rx_bps': 0, 'tx_bps': 0}
    pool.disconnect()
    return jsonify(resultados)

# Endpoint: CPU y memoria
@app.route('/api/cpu_memoria')
def api_cpu_memoria():
    pool, api = conectar_api()
    res = api.get_resource('/system/resource').get()[0]
    pool.disconnect()

    cpu = float(res.get('cpu-load', 0))
    mem_free = int(res.get('free-memory', 0))
    mem_total = int(res.get('total-memory', 1))
    mem_used_pct = 100 * (mem_total - mem_free) / mem_total if mem_total else 0

    return jsonify({'cpu': cpu, 'memoria': mem_used_pct})

# Endpoint: firewall resumen
@app.route('/api/firewall')
def api_firewall():
    pool, api = conectar_api()
    reglas = api.get_resource('/ip/firewall/filter').get()
    pool.disconnect()
    activas = sum(1 for r in reglas if not r.get('disabled', False))
    desactivadas = sum(1 for r in reglas if r.get('disabled', False))
    return jsonify({'activas': activas, 'desactivadas': desactivadas})

# Endpoint: errores y paquetes perdidos por interfaz
from flask import jsonify

# Endpoint: latencia ping (simulado, cambia con método real si quieres)
import random
@app.route('/api/ping')
def api_ping():
    # Aquí puedes hacer ping real desde backend o simular
    latencias = [random.randint(20, 80) for _ in range(30)]
    return jsonify(latencias)

import platform

@app.route('/tools_ping', methods=['GET', 'POST'])
def tool_ping():
    resultado = None
    if request.method == 'POST':
        host = request.form.get('host')
        if host:
            try:
                # Detectar sistema operativo
                param = '-n' if platform.system().lower() == 'windows' else '-c'
                result = subprocess.run(["ping", param, "4", host], capture_output=True, text=True, timeout=10)
                resultado = result.stdout
            except Exception as e:
                resultado = f"Error: {e}"
        else:
            resultado = "Debe ingresar un host"
    return render_template('tools_ping.html', resultado=resultado)



# PORTSCAN
@app.route('/tools/portscan', methods=['GET', 'POST'])
def tool_portscan():
    results = []
    host = ""
    ports = "21,22,23,25,53,80,110,143,443,8080"
    warning = None

    if request.method == 'POST':
        host = request.form.get('host')
        ports_input = request.form.get('ports', '21,22,23,25,53,80,110,143,443,8080')
        ports = ports_input

        try:
            port_list = [int(p.strip()) for p in ports_input.split(',') if p.strip().isdigit()]
        except:
            port_list = [21,22,23,25,53,80,110,143,443,8080]
            warning = "Error al interpretar los puertos, usando los predeterminados."

        if not host:
            warning = "Debes seleccionar una IP válida."

        for port in port_list:
            try:
                with socket.create_connection((host, port), timeout=1):
                    results.append({"port": port, "status": "open"})
            except:
                results.append({"port": port, "status": "closed"})

    # Aquí obtienes las IPs del Mikrotik (asegúrate que funcione esto)
    mikrotik_ips = obtener_todas_las_ips_mikrotik()

    return render_template("tools_portscan.html", ips=mikrotik_ips, results=results, host=host, ports=ports, warning=warning)

def ip_alive(host):
    param = "-n" if platform.system().lower() == "windows" else "-c"
    response = subprocess.run(["ping", param, "1", host], stdout=subprocess.DEVNULL)
    return response.returncode == 0

@app.context_processor
def utility_processor():
    def get_port_description(port):
        descriptions = {
            21: "FTP: Transferencia de archivos.",
            22: "SSH: Acceso remoto seguro.",
            23: "Telnet: Acceso remoto no seguro.",
            25: "SMTP: Envío de correos.",
            53: "DNS: Resolución de dominios.",
            80: "HTTP: Navegación web estándar.",
            110: "POP3: Lectura de correos.",
            143: "IMAP: Sincronización de correos.",
            443: "HTTPS: Navegación web segura.",
            8080: "HTTP Alternativo: Proxy u otros servicios web."
        }
        return descriptions.get(port, "Puerto común o personalizado.")
    return dict(get_port_description=get_port_description)


# GEOIP
@app.route('/tools/geoip', methods=['POST'])
def tool_geoip():
    ip = request.json.get('ip')
    if not ip:
        return jsonify({"error": "Falta la IP"}), 400
    try:
        res = requests.get(f"https://ipinfo.io/{ip}/json", timeout=5)
        data = res.json()

        # Detectar bogon o IP privada
        if data.get("bogon", False):
            return jsonify({
                "error": "La IP consultada es una IP privada o no enrutada (bogon).",
                "ip": ip
            }), 200
        
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/tools/geoip', methods=['GET'])
def tool_geoip_page():
    # Aquí renderizas la plantilla HTML
    return render_template('tools_geoip.html')

# DNS LOOKUP
@app.route('/tools/dnslookup', methods=['GET', 'POST'])
def tool_dns_lookup():
    if request.method == 'POST':
        domain = (request.json or request.form).get('domain')
        if not domain:
            return jsonify({"error": "Dominio requerido"}), 400
        try:
            answers = dns.resolver.resolve(domain, 'A')
            ips = [r.to_text() for r in answers]
            if request.is_json:
                return jsonify({"domain": domain, "ips": ips})
            else:
                return render_template('tools_dnslookup.html', domain=domain, ips=ips)
        except Exception as e:
            error = {"error": str(e)}
            if request.is_json:
                return jsonify(error), 500
            else:
                return render_template('tools_dnslookup.html', error=error)
    return render_template('tools_dnslookup.html')

@app.route('/tools/whois_page')
def whois_page():
    return render_template('tools_whois.html')

def clean_whois_data(data):
    def clean_value(val):
        if isinstance(val, list):
            return ', '.join(clean_value(v) for v in val)
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d %H:%M:%S')
        if val is None:
            return ""
        return str(val)
    return {k: clean_value(v) for k, v in data.items()}

@app.route('/tools/whois', methods=['POST'])
def tool_whois():
    domain = request.json.get('domain')
    if not domain:
        return jsonify({"error": "Dominio requerido"}), 400
    try:
        info = whois.whois(domain)
        clean_info = clean_whois_data(info)
        return jsonify(clean_info)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


from routeros_api import RouterOsApiPool

def obtener_interfaces(api):
    # Obtiene todas las interfaces y su estado "running"
    interfaces = api.get_resource('/interface').get()
    return interfaces

def obtener_clientes_pppoe(api):
    # Obtiene clientes PPP activos
    ppp_active = api.get_resource('/ppp/active').get()
    return ppp_active

def obtener_clientes_hotspot(api):
    # Obtiene clientes Hotspot activos
    hotspot_active = api.get_resource('/ip/hotspot/active').get()
    return hotspot_active

@app.route('/api/clientes_conectados')
def api_clientes_conectados():
    pool, api = conectar_api()
    try:
        pppoe_activos = len(obtener_clientes_pppoe(api))
        hotspot_activos = len(obtener_clientes_hotspot(api))
    finally:
        pool.disconnect()
    return jsonify({'pppoe': pppoe_activos, 'hotspot': hotspot_activos})

@app.route('/api/top5_clientes_trafico')
def api_top5_clientes_trafico():
    pool, api = conectar_api()

    # Obtener clientes PPPoE activos con su tráfico
    clientes_pppoe = api.get_resource('/ppp/active').get()
    monitor = api.get_resource('/interface/monitor-traffic')

    clientes_trafico = []

    for cliente in clientes_pppoe:
        iface = cliente.get('interface')
        nombre = cliente.get('caller-id') or cliente.get('name') or 'Desconocido'

        try:
            data = monitor.call('monitor', {'interface': iface, 'once': True})
            rx_bps = int(data[0].get('rx-bits-per-second', 0)) if data else 0
            tx_bps = int(data[0].get('tx-bits-per-second', 0)) if data else 0
        except:
            rx_bps = 0
            tx_bps = 0

        total_bps = rx_bps + tx_bps
        clientes_trafico.append({'cliente': nombre, 'trafico_bps': total_bps})

    # Ordenar y tomar top 5
    top5 = sorted(clientes_trafico, key=lambda x: x['trafico_bps'], reverse=True)[:5]

    pool.disconnect()
    return jsonify(top5)


def obtener_reglas_firewall(api):
    # Obtiene todas las reglas de firewall
    reglas = api.get_resource('/ip/firewall/filter').get()
    return reglas

def obtener_logs(api, limite=10):
    logs = api.get_resource('/log').get()
    return logs[-limite:]  # Obtener solo los últimos 'limite' registros

def obtener_trafico_todas_interfaces(api):
    interfaces_res = api.get_resource('/interface')
    interfaces = interfaces_res.get()
    monitor = api.get_resource('/interface/monitor-traffic')

    resultados = {}

    # Filtrar solo interfaces "running" y que sean de tipo física o VLAN
    interfaces_filtradas = [iface for iface in interfaces if iface.get('running') and iface.get('type') in ['ether', 'wireless', 'vlan']]

    for iface in interfaces_filtradas:
        name = iface.get('name')
        if not name:
            continue
        try:
            data = monitor.call('monitor', {'interface': name, 'once': True})
            if data and len(data) > 0:
                rx_bps = int(data[0].get('rx-bits-per-second', 0))
                tx_bps = int(data[0].get('tx-bits-per-second', 0))
                resultados[name] = {'rx_bps': rx_bps, 'tx_bps': tx_bps}
            else:
                resultados[name] = {'rx_bps': 0, 'tx_bps': 0}
        except Exception:
            resultados[name] = {'rx_bps': 0, 'tx_bps': 0}
    return resultados

def obtener_ips_usadas():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT ip_asignada FROM contratos")
    usadas = [row[0] for row in cur.fetchall()]
    conn.close()
    return usadas

def obtener_ip_disponible():
    ip_base = '192.168.69.'
    usadas = obtener_ips_usadas()
    for i in range(100, 200):
        ip = f"{ip_base}{i}"
        if ip not in usadas:
            return ip
    return None

def generar_numero_contrato():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT numero_contrato FROM contratos ORDER BY id DESC LIMIT 1")
    ultimo = cur.fetchone()
    conn.close()

    if ultimo and ultimo[0]:  # Si hay al menos un contrato y no es None
        try:
            ultimo_num = int(ultimo[0].split('-')[1])
        except (IndexError, ValueError):
            ultimo_num = 0
    else:
        ultimo_num = 0

    nuevo_num = ultimo_num + 1
    return f'CT-{nuevo_num:04d}'  # Por ejemplo: CON-0001, CON-0002, etc.

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Asumiendo plan-límites predefinidos:
PLAN_LIMITES = {
    '5 Mbps':  {'upload': '5000000',  'download': '5000000'},
    '10 Mbps': {'upload': '10000000', 'download': '10000000'},
    '15 Mbps': {'upload': '15000000', 'download': '15000000'},
    '20 Mbps': {'upload': '20000000', 'download': '20000000'},
    '25 Mbps': {'upload': '25000000', 'download': '25000000'},
    '30 Mbps': {'upload': '30000000', 'download': '30000000'}
}

PLAN_PRECIOS = {
    '5 Mbps': 20,
    '10 Mbps': 35,
    '15 Mbps': 50,
    '20 Mbps': 65,
    '25 Mbps': 80,
    '30 Mbps': 100
}

@app.route('/contratos', methods=['GET'])
def ver_formulario_contratos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM contratos ORDER BY id DESC")
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    contratos = [dict(zip(columnas, fila)) for fila in filas]
    conn.close()

    # Aquí conectamos a Mikrotik para obtener el estado de las IPs
    try:
        pool, api = conectar_api()
        resource_queues = api.get_resource('/queue/simple')
        queues = resource_queues.get()

        resource_addrlist = api.get_resource('/ip/firewall/address-list')
        cortados = resource_addrlist.get(list='bloqueados')
        ips_cortadas = {c['address'] for c in cortados}

        # Crear diccionario IP -> estado
        estados_por_ip = {}
        for q in queues:
            ip = q.get('target', '').split('/')[0]
            if q.get('disabled') == 'true':
                estado = 'Inactivo'
            elif ip in ips_cortadas:
                estado = 'Cortado'
            else:
                estado = 'Activo'
            estados_por_ip[ip] = estado

        pool.disconnect()

        # Asignar estado a cada contrato según su IP
        for c in contratos:
            c['estado'] = estados_por_ip.get(c['ip_asignada'], 'Activo')

    except Exception as e:
        # Si hay error, asignar estado activo por defecto y loguear error si quieres
        for c in contratos:
            c['estado'] = 'Activo'
        print(f"Error al consultar estado en Mikrotik: {e}")

    ip_disponible = obtener_ip_disponible()
    numero_contrato = generar_numero_contrato()

    return render_template('contratos.html',
                           contratos=contratos,
                           ip_disponible=ip_disponible,
                           numero_contrato=numero_contrato,
                           provincias_list=provincias_list)


@app.route('/contratos/crear', methods=['POST'])
def crear_contrato():
    data = request.form
    imagen = request.files.get('imagen_domicilio')
    filename = None

    # Guardar imagen si se adjunta
    if imagen and imagen.filename:
        filename = secure_filename(imagen.filename)
        imagen.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    # Fechas automáticas
    fecha_inicio = date.today()

    hoy = date.today()
    if hoy.day <= 5:
        fecha_corte = date(hoy.year, hoy.month, 5)
    else:
        mes = hoy.month + 1 if hoy.month < 12 else 1
        año = hoy.year if mes != 1 else hoy.year + 1
        fecha_corte = date(año, mes, 5)

    # Fecha de fin y suspensión se inicializan como None
    fecha_fin = None
    fecha_suspension = None

    # Datos automáticos
    numero_contrato = generar_numero_contrato()
    ip = data['ip_asignada']
    plan = data['plan_internet']
    limites = PLAN_LIMITES.get(plan, {'upload': '1000000', 'download': '1000000'})

    # Guardar en base de datos
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO contratos (
            numero_contrato, nombres, apellidos, cedula, direccion, telefono, email,
            pais, provincia, canton, latitud, longitud,
            plan_internet, ip_asignada, mascara, gateway,
            imagen_domicilio, fecha_corte, fecha_inicio, fecha_fin, fecha_suspension
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        numero_contrato,
        data.get('nombres'), data.get('apellidos'), data.get('cedula'),
        data.get('direccion'), data.get('telefono'), data.get('email'),
        data.get('pais'), data.get('provincia'), data.get('canton'),
        data.get('latitud'), data.get('longitud'),
        plan, ip, data.get('mascara'), data.get('gateway'),
        filename, fecha_corte, fecha_inicio, fecha_fin, fecha_suspension
    ))

    conn.commit()
    conn.close()

    # Agregar queue en MikroTik
    pool, api = conectar_api()
    q = api.get_resource('/queue/simple')
    q.add(
        name=numero_contrato,
        target=f"{ip}/32",
        **{'max-limit': f"{limites['upload']}/{limites['download']}"}
    )
    pool.disconnect()

    return redirect('/contratos')

@app.route('/contratos/toggle_estado/<int:id>', methods=['POST'])
def toggle_estado(id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT estado, ip_asignada FROM contratos WHERE id = %s", (id,))
    contrato = cur.fetchone()

    if contrato:
        estado_actual, ip = contrato
        hoy = date.today()

        if estado_actual == 'Activo':
            # 1. Intentar cortar en MikroTik
            response = requests.post(f"{request.host_url}api/queues/cortar", json={"ip": ip})
            if response.status_code == 200 and response.json().get("status") == "ok":
                nuevo_estado = 'Cortado'
                cur.execute("UPDATE contratos SET estado = %s, fecha_corte = %s WHERE id = %s",
                            (nuevo_estado, hoy, id))
        else:
            # 2. Intentar reconectar en MikroTik
            response = requests.post(f"{request.host_url}api/queues/reconectar", json={"ip": ip})
            if response.status_code == 200 and response.json().get("status") == "ok":
                nuevo_estado = 'Activo'
                cur.execute("UPDATE contratos SET estado = %s, fecha_corte = NULL WHERE id = %s",
                            (nuevo_estado, id))

        conn.commit()
    conn.close()
    return redirect('/contratos')


@app.route('/pago', methods=['GET', 'POST'])
def pagar_contrato():
    conn = get_db_connection()
    cur = conn.cursor()
    comprobante_filename = None

    if request.method == 'POST':
        contrato_id = request.form['contrato_id']
        cliente = request.form['cliente']
        monto = request.form['monto']
        metodo_pago = request.form['metodo_pago']
        fecha = datetime.now()

        # Obtener plan actual del contrato
        cur.execute("""
            SELECT plan_internet
            FROM contratos
            WHERE id = %s
        """, (contrato_id,))
        contrato_info = cur.fetchone()

        if not contrato_info:
            flash('Contrato no encontrado', 'error')
            return redirect(url_for('pagar_contrato'))

        plan_internet_guardado = contrato_info[0]
        # Usar el diccionario de precios
        precio_plan_guardado = PLAN_PRECIOS.get(plan_internet_guardado, 0)

        # Si es transferencia, guardar comprobante
        if metodo_pago == 'Transferencia':
            file = request.files.get('comprobante')
            if file and allowed_file(file.filename):
                filename = secure_filename(f"{contrato_id}_{fecha.strftime('%Y%m%d%H%M%S')}_{file.filename}")
                base_dir = os.path.dirname(os.path.abspath(__file__))
                comprobante_path = os.path.join(base_dir, 'static', 'comprobantes')
                os.makedirs(comprobante_path, exist_ok=True)
                file.save(os.path.join(comprobante_path, filename))
                comprobante_filename = filename
            else:
                flash('Debe subir un archivo válido (pdf, jpg, png)', 'error')
                return redirect(url_for('pagar_contrato'))

        # Guardar el pago con plan y precio actual
        cur.execute("""
            INSERT INTO pagos (contrato_id, cliente, monto, metodo_pago, fecha, comprobante, tipo_recibo,
                               plan_internet_guardado, precio_plan_guardado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (contrato_id, cliente, monto, metodo_pago, fecha, comprobante_filename, metodo_pago,
              plan_internet_guardado, precio_plan_guardado))
        conn.commit()

        # Verificar si se debe reconectar el contrato
        cur.execute("SELECT estado, ip_asignada FROM contratos WHERE id = %s", (contrato_id,))
        contrato = cur.fetchone()

        if contrato:
            estado, ip = contrato
            flash(f"Estado actual del contrato: '{estado}'", 'info')

            if estado and estado.strip().lower() == 'cortado':
                try:
                    response = requests.post(f"{request.host_url}api/queues/reconectar", json={"ip": ip})
                    if response.status_code == 200 and response.json().get("status") == "ok":
                        cur.execute(
                            "UPDATE contratos SET estado = %s, fecha_corte = NULL WHERE id = %s",
                            ('Activo', contrato_id)
                        )
                        conn.commit()
                        flash('Pago registrado y contrato reconectado correctamente', 'success')
                    else:
                        flash('Pago registrado, pero no se pudo reconectar el contrato automáticamente', 'warning')
                except Exception as e:
                    flash(f'Pago registrado, pero error al intentar reconectar: {str(e)}', 'warning')
            else:
                flash('Pago registrado correctamente', 'success')
        else:
            flash('Pago registrado correctamente', 'success')

        return redirect(url_for('pagar_contrato'))

    # Datos de contratos para el formulario
    cur.execute("""
        SELECT id, numero_contrato, nombres, apellidos, plan_internet, fecha_corte
        FROM contratos ORDER BY nombres ASC
    """)
    contratos = cur.fetchall()

    # Datos de pagos para historial
    cur.execute("""
        SELECT pagos.id, pagos.cliente, pagos.monto, pagos.metodo_pago, pagos.fecha, pagos.comprobante, pagos.tipo_recibo,
               pagos.plan_internet_guardado, pagos.precio_plan_guardado,
               c.numero_contrato, c.nombres, c.apellidos, c.plan_internet, c.fecha_corte, c.estado
        FROM pagos
        JOIN contratos c ON pagos.contrato_id = c.id
        ORDER BY pagos.fecha DESC
    """)
    pagos = cur.fetchall()

    conn.close()
    return render_template("pago.html", contratos=contratos, pagos=pagos, plan_precios=PLAN_PRECIOS)



@app.route('/resumen_financiero')
def resumen_financiero():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Obtener pagos con la información necesaria
    cur.execute("""
    SELECT 
        pagos.fecha,
        pagos.cliente,
        pagos.plan_internet_guardado,
        pagos.monto,
        pagos.metodo_pago,
        contratos.numero_contrato,
        contratos.estado
    FROM pagos
    LEFT JOIN contratos ON pagos.contrato_id = contratos.id
    ORDER BY pagos.fecha DESC
    """)
    pagos = cur.fetchall()

    # Totales
    cur.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE fecha::date = CURRENT_DATE")
    total_diario = cur.fetchone()['coalesce']
    
    cur.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE DATE_TRUNC('month', fecha) = DATE_TRUNC('month', CURRENT_DATE)")
    total_mensual = cur.fetchone()['coalesce']

    cur.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE DATE_TRUNC('year', fecha) = DATE_TRUNC('year', CURRENT_DATE)")
    total_anual = cur.fetchone()['coalesce']

    cur.execute("SELECT COALESCE(SUM(monto),0) FROM pagos")
    total_general = cur.fetchone()['coalesce']

    # Datos para gráfico mensual (suma por mes del año actual)
    cur.execute("""
        SELECT EXTRACT(MONTH FROM fecha) AS mes, COALESCE(SUM(monto),0) AS total
        FROM pagos
        WHERE EXTRACT(YEAR FROM fecha) = EXTRACT(YEAR FROM CURRENT_DATE)
        GROUP BY mes
        ORDER BY mes
    """)
    mensual_data = cur.fetchall()
    labels = [int(row['mes']) for row in mensual_data]
    valores = [float(row['total']) for row in mensual_data]

    # Datos para gráfico anual (suma por año)
    cur.execute("""
        SELECT EXTRACT(YEAR FROM fecha) AS anio, COALESCE(SUM(monto),0) AS total
        FROM pagos
        GROUP BY anio
        ORDER BY anio
    """)
    anual_data = cur.fetchall()
    labels_anual = [int(row['anio']) for row in anual_data]
    valores_anual = [float(row['total']) for row in anual_data]

    # Datos para gráfico por método de pago
    cur.execute("""
        SELECT metodo_pago, COALESCE(SUM(monto),0) AS total
        FROM pagos
        GROUP BY metodo_pago
        ORDER BY metodo_pago
    """)
    metodos_data = cur.fetchall()
    metodos = []
    valores_metodos = []
    for row in metodos_data:
        metodo = row['metodo_pago'] if row['metodo_pago'] else 'No especificado'
        total = row['total'] if row['total'] is not None else 0
        metodos.append(metodo)
        valores_metodos.append(float(total))

    # Datos para gráfico por plan de internet
    cur.execute("""
        SELECT plan_internet_guardado, COALESCE(SUM(monto),0) AS total
        FROM pagos
        GROUP BY plan_internet_guardado
        ORDER BY plan_internet_guardado
    """)
    planes_data = cur.fetchall()
    planes = []
    valores_planes = []
    for row in planes_data:
        plan = row['plan_internet_guardado'] if row['plan_internet_guardado'] else 'No especificado'
        total = row['total'] if row['total'] is not None else 0
        planes.append(plan)
        valores_planes.append(float(total))

    cur.close()
    conn.close()

    return render_template('resumen_financiero.html',
                           total_diario=round(total_diario, 2),
                           total_mensual=round(total_mensual, 2),
                           total_anual=round(total_anual, 2),
                           total_general=round(total_general, 2),
                           pagos=pagos,
                           labels=labels,
                           valores=valores,
                           labels_anual=labels_anual,
                           valores_anual=valores_anual,
                           metodos=metodos,
                           valores_metodos=valores_metodos,
                           planes=planes,
                           valores_planes=valores_planes)

@app.route('/cuentas_por_cobrar')
def cuentas_por_cobrar():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT contratos.numero_contrato, contratos.nombres, contratos.apellidos,
               COALESCE(SUM(pagos.monto), 0) AS pagado,
               contratos.plan_internet, contratos.estado
        FROM contratos
        LEFT JOIN pagos ON pagos.contrato_id = contratos.id
        WHERE contratos.estado = 'Activo'
        GROUP BY contratos.id
    """)
    cuentas = cur.fetchall()
    
    cur.close()
    conn.close()

    return render_template('cuentas_por_cobrar.html', cuentas=cuentas)

@app.route('/cuentas_por_pagar')
def cuentas_por_pagar():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT proveedor, descripcion, monto, fecha_vencimiento, estado
        FROM cuentas_por_pagar
        ORDER BY fecha_vencimiento ASC
    """)
    cuentas = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('cuentas_por_pagar.html', cuentas=cuentas)

@app.route('/cuentas_por_pagar/nueva', methods=['POST'])
def nueva_cuenta_por_pagar():
    proveedor = request.form['proveedor']
    descripcion = request.form['descripcion']
    monto = request.form['monto']
    fecha_vencimiento = request.form['fecha_vencimiento']
    estado = request.form['estado']

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO cuentas_por_pagar (proveedor, descripcion, monto, fecha_vencimiento, estado)
        VALUES (%s, %s, %s, %s, %s)
    """, (proveedor, descripcion, monto, fecha_vencimiento, estado))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('cuentas_por_pagar'))


@app.route('/caja_bancos')
def caja_bancos():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT id, tipo, nombre_cuenta, saldo_actual, descripcion, fecha_creacion
        FROM cuentas_bancarias
        ORDER BY fecha_creacion DESC
    """)
    cuentas = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template('caja_bancos.html', cuentas=cuentas)


@app.route('/caja_bancos/nueva', methods=['POST'])
def nueva_cuenta_bancaria():
    tipo = request.form['tipo']
    nombre_cuenta = request.form['nombre_cuenta']
    saldo_actual = request.form['saldo_actual']
    descripcion = request.form.get('descripcion')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO cuentas_bancarias (tipo, nombre_cuenta, saldo_actual, descripcion)
        VALUES (%s, %s, %s, %s)
    """, (tipo, nombre_cuenta, saldo_actual, descripcion))
    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('caja_bancos'))


@app.route('/presupuestos')
def presupuestos():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT categoria, monto_estimado, monto_real, periodo
        FROM presupuestos
        ORDER BY periodo DESC
    """)
    datos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('presupuestos.html', presupuestos=datos)

@app.route('/gastos')
def gastos_generales():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT fecha, descripcion, monto, categoria
        FROM gastos
        ORDER BY fecha DESC
    """)
    gastos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('gastos.html', gastos=gastos)

@app.route('/reportes_financieros')
def reportes_financieros():
    # Aquí podrías incluir links para descargar PDF/Excel o ver reportes consolidados
    return render_template('reportes_financieros.html')

@app.route('/gestion_impuestos')
def gestion_impuestos():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT tipo_impuesto, monto, fecha_pago, periodo, estado
        FROM impuestos
        ORDER BY fecha_pago DESC
    """)
    impuestos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('gestion_impuestos.html', impuestos=impuestos)


@app.route('/recibo/<int:pago_id>')
def imprimir_recibo(pago_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT pagos.id, pagos.cliente, pagos.monto, pagos.metodo_pago, pagos.fecha,
               c.numero_contrato, c.plan_internet, c.fecha_corte
        FROM pagos
        JOIN contratos c ON pagos.contrato_id = c.id
        WHERE pagos.id = %s
    """, (pago_id,))
    pago = cur.fetchone()
    conn.close()

    if pago is None:
        return "Recibo no encontrado", 404

    return render_template('recibo.html', pago=pago)

@app.route('/pago/eliminar/<int:pago_id>', methods=['POST'])
def eliminar_pago(pago_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Traer nombre del archivo del comprobante para borrarlo si existe
    cur.execute("SELECT comprobante FROM pagos WHERE id = %s", (pago_id,))
    result = cur.fetchone()

    if result and result[0]:  # Si tiene archivo comprobante
        comprobante_path = os.path.join(current_app.root_path, 'static', 'comprobantes', result[0])
        if os.path.exists(comprobante_path):
            os.remove(comprobante_path)

    # Eliminar el pago
    cur.execute("DELETE FROM pagos WHERE id = %s", (pago_id,))
    conn.commit()
    conn.close()

    flash("Pago eliminado correctamente", "success")
    return redirect(url_for('pagar_contrato'))


@app.route('/contratos/suspender/<int:id>', methods=['POST'])
def suspender_contrato(id):
    hoy = date.today()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT ip_asignada FROM contratos WHERE id = %s", (id,))
    fila = cur.fetchone()
    if not fila:
        conn.close()
        return "Contrato no encontrado", 404

    ip = fila[0]

    # Cortar acceso en MikroTik al suspender
    response = requests.post(f"{request.host_url}api/queues/cortar", json={"ip": ip})
    if response.status_code == 200 and response.json().get("status") == "ok":
        cur.execute("""
            UPDATE contratos
            SET estado = 'Suspendido', fecha_suspension = %s
            WHERE id = %s
        """, (hoy, id))
        conn.commit()

    conn.close()
    return redirect('/contratos')



@app.route('/contratos/cortar_reconectar/<int:id>', methods=['POST'])
def cortar_reconectar_contrato(id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener IP y estado actual del contrato
    cur.execute("SELECT ip_asignada, estado FROM contratos WHERE id = %s", (id,))
    fila = cur.fetchone()
    if not fila:
        conn.close()
        return "Contrato no encontrado", 404

    ip_asignada, estado_actual = fila

    # Definir URL base de la API
    base_api_url = request.host_url.rstrip('/')

    try:
        if estado_actual == 'Activo':
            # Enviar solicitud para cortar
            resp = requests.post(f"{base_api_url}/api/queues/cortar", json={'ip': ip_asignada})
            if resp.status_code == 200 and resp.json().get('status') == 'ok':
                nuevo_estado = 'Cortado'
                fecha_corte = date.today()
            else:
                print(f"[ERROR cortando Mikrotik] Código: {resp.status_code}, Respuesta: {resp.text}")
                conn.close()
                return f"Error cortando en MikroTik: {resp.text}", 500

        else:
            # Enviar solicitud para reconectar
            resp = requests.post(f"{base_api_url}/api/queues/reconectar", json={'ip': ip_asignada})
            if resp.status_code == 200 and resp.json().get('status') == 'ok':
                nuevo_estado = 'Activo'
                fecha_corte = None
            else:
                print(f"[ERROR reconectando Mikrotik] Código: {resp.status_code}, Respuesta: {resp.text}")
                conn.close()
                return f"Error reconectando en MikroTik: {resp.text}", 500

        # Actualizar el estado del contrato solo si la acción fue exitosa en MikroTik
        cur.execute("""
            UPDATE contratos
            SET estado = %s,
                fecha_corte = %s
            WHERE id = %s
        """, (nuevo_estado, fecha_corte, id))
        conn.commit()

    except Exception as e:
        print(f"[ERROR general en cortar_reconectar_contrato] {e}")
        return f"Error inesperado: {str(e)}", 500
    finally:
        conn.close()

    return redirect('/contratos')


def editar_simple_queue(nombre, nueva_ip, nuevo_plan):
    plan_to_limit = {
        "1M": "1M/1M",
        "5M": "5M/5M",
        "10M": "10M/10M",
        "15M": "15M/15M",
        "20M": "20M/20M",
        "25M": "25M/25M",
        "30M": "30M/30M",
        "Sin límite": "0/0"
    }

    try:
        pool, api = conectar_api()
        queues = api.get_resource('/queue/simple')

        for q in queues.get():
            if q.get('name') == nombre:
                queue_id = q.get('.id') or q.get('id')
                if not queue_id:
                    print("No se encontró ID para la queue:", nombre)
                    break

                max_limit = plan_to_limit.get(nuevo_plan.strip(), "0/0")
                queues.set(
                    id=queue_id,
                    target=f"{nueva_ip}/32",
                    max_limit=max_limit
                )
                print(f"Queue actualizada: {nombre}")
                break
        else:
            print(f"No se encontró la Simple Queue con el nombre: {nombre}")

    except Exception as e:
        print("Error al editar Simple Queue:", e)

    finally:
        pool.disconnect()


@app.route('/contratos/editar/<int:id>', methods=['GET', 'POST'])
def editar_contrato(id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        data = request.form
        imagen = request.files.get('imagen_domicilio')
        filename = None

        # Obtener el nombre anterior para no perderlo si no hay nueva imagen
        cur.execute("SELECT imagen_domicilio FROM contratos WHERE id = %s", (id,))
        fila = cur.fetchone()
        imagen_anterior = fila[0] if fila else None

        if imagen:
            filename = secure_filename(imagen.filename)
            imagen.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        else:
            filename = imagen_anterior

        cur.execute("""
            UPDATE contratos SET
                numero_contrato = %s,
                nombres = %s,
                apellidos = %s,
                cedula = %s,
                direccion = %s,
                telefono = %s,
                email = %s,
                pais = %s,
                provincia = %s,
                canton = %s,
                latitud = %s,
                longitud = %s,
                plan_internet = %s,
                ip_asignada = %s,
                mascara = %s,
                gateway = %s,
                imagen_domicilio = %s,
                fecha_corte = %s,
                fecha_inicio = %s,
                fecha_fin = %s,
                fecha_suspension = %s
            WHERE id = %s
        """, (
            data.get('numero_contrato'), data.get('nombres'), data.get('apellidos'),
            data.get('cedula'), data.get('direccion'), data.get('telefono'), data.get('email'),
            data.get('pais'), data.get('provincia'), data.get('canton'),
            data.get('latitud'), data.get('longitud'), data.get('plan_internet'),
            data.get('ip_asignada'), data.get('mascara'), data.get('gateway'),
            filename,
            data.get('fecha_corte') or None,
            data.get('fecha_inicio') or None,
            data.get('fecha_fin') or None,
            data.get('fecha_suspension') or None,
            id
        ))
        conn.commit()

        # Actualizar en Mikrotik
        editar_simple_queue(data.get('numero_contrato'), data.get('ip_asignada'), data.get('plan_internet'))

        conn.close()
        return redirect('/contratos')

    else:
        cur.execute("SELECT * FROM contratos WHERE id = %s", (id,))
        columnas = [desc[0] for desc in cur.description]
        fila = cur.fetchone()
        contrato = dict(zip(columnas, fila)) if fila else {}
        conn.close()
        return render_template('editar_contrato.html', contrato=contrato)

@app.route('/contratos/eliminar/<int:id>', methods=['POST'])
def eliminar_contrato(id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener la IP asociada al contrato
    cur.execute("SELECT ip_asignada FROM contratos WHERE id = %s", (id,))
    fila = cur.fetchone()
    if not fila:
        conn.close()
        return "Contrato no encontrado", 404

    ip_asignada = fila[0]

    try:
        # Conexión a MikroTik
        pool, api = conectar_api()

        # 1. Eliminar de address-list si está presente
        addr_list = api.get_resource('/ip/firewall/address-list')
        registros_addr = addr_list.get()
        for r in registros_addr:
            if r.get('address') == ip_asignada:
                addr_list.remove(id=r['id'])
                print(f"[INFO] IP {ip_asignada} eliminada de address-list")

        # 2. Eliminar queue simple si existe
        queues = api.get_resource('/queue/simple')
        registros_queue = queues.get()
        for q in registros_queue:
            if q.get('target') and ip_asignada in q.get('target'):
                queues.remove(id=q['id'])
                print(f"[INFO] Queue para IP {ip_asignada} eliminada")

        pool.disconnect()

    except Exception as e:
        print(f"[ERROR MikroTik] {e}")
        conn.close()
        return f"Error eliminando en MikroTik: {e}", 500

    # 3. Eliminar el contrato de la base de datos
    cur.execute("DELETE FROM contratos WHERE id = %s", (id,))
    conn.commit()
    conn.close()

    print(f"[INFO] Contrato {id} eliminado exitosamente.")
    return redirect('/contratos')


@app.route('/clientes')
def clientes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM contratos ORDER BY id DESC")
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    contratos = [dict(zip(columnas, fila)) for fila in filas]
    conn.close()
    return render_template('clientes.html', contratos=contratos)

@app.route('/clientes/<int:id>')
def ver_contrato(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM contratos WHERE id = %s", (id,))
    fila = cur.fetchone()
    if not fila:
        conn.close()
        return "Contrato no encontrado", 404

    columnas = [desc[0] for desc in cur.description]
    contrato = dict(zip(columnas, fila))
    conn.close()
    return render_template('ver_contrato.html', contrato=contrato)


def obtener_contratos_con_estado():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM contratos ORDER BY id DESC")
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    contratos = [dict(zip(columnas, fila)) for fila in filas]
    conn.close()

    try:
        pool, api = conectar_api()
        resource_queues = api.get_resource('/queue/simple')
        queues = resource_queues.get()

        resource_addrlist = api.get_resource('/ip/firewall/address-list')
        cortados = resource_addrlist.get(list='bloqueados')
        ips_cortadas = {c['address'] for c in cortados}

        estados_por_ip = {}
        for q in queues:
            ip = q.get('target', '').split('/')[0]
            if q.get('disabled') == 'true':
                estado = 'Inactivo'
            elif ip in ips_cortadas:
                estado = 'Cortado'
            else:
                estado = 'Activo'
            estados_por_ip[ip] = estado

        pool.disconnect()

        for c in contratos:
            c['estado'] = estados_por_ip.get(c['ip_asignada'], 'Activo')

    except Exception as e:
        for c in contratos:
            c['estado'] = 'Activo'
        print(f"Error al consultar estado en Mikrotik: {e}")

    return contratos

@app.route('/dhcp')
def dhcp():
    leases = mikrotik_execute('/ip/dhcp-server/lease', 'list')
    if 'error' in leases:
        return f"Error al obtener los leases DHCP: {leases['error']}"

    # Agrupar por servidor DHCP si está disponible
    grouped = {}
    for lease in leases:
        server = lease.get('server', 'Sin definir')
        grouped.setdefault(server, []).append(lease)

    return render_template('dhcp.html', leases_grouped=grouped)

@app.route('/dhcp/delete/<lease_id>', methods=['POST'])
def delete_lease(lease_id):
    result = mikrotik_execute('/ip/dhcp-server/lease', 'remove', {'id': lease_id})
    if 'error' in result:
        return jsonify({'success': False, 'error': result['error']})
    return jsonify({'success': True})

@app.route('/api/dhcp', methods=['GET'])
def api_dhcp():
    leases = mikrotik_execute('/ip/dhcp-server/lease', 'list')
    if 'error' in leases:
        return jsonify({'error': leases['error']}), 500

    return jsonify(leases)


@app.route('/dhcp')
def lista_clientes():
    clientes = []
    bloqueados_ips = []
    error = None

    try:
        pool, api = conectar_api()

        leases = api.get_resource('/ip/dhcp-server/lease').get()
        bloqueados = api.get_resource('/ip/firewall/address-list').get()
        bloqueados_ips = []
        for b in bloqueados:
         if b.get('list') == 'bloqueados' and 'address' in b:
          bloqueados_ips.append(b['address'])

        pool.disconnect()
        clientes = leases
    except Exception as e:
        error = f'Error: {str(e)}'

    return render_template('clientes.html', clientes=clientes, bloqueados=bloqueados_ips, error=error)

@app.route('/cortar/<ip>')
def cortar_cliente(ip):
    try:
        pool, api = conectar_api()
        lista = api.get_resource('/ip/firewall/address-list')
        lista.add(address=ip, list='bloqueados', comment='Corte automático')
        pool.disconnect()
    except:
        pass
    return redirect(url_for('lista_clientes'))

@app.route('/reconectar/<ip>')
def reconectar_cliente(ip):
    try:
        pool, api = conectar_api()
        lista = api.get_resource('/ip/firewall/address-list')

        # Buscar entradas que coincidan exactamente
        registros = lista.get()
        for r in registros:
            if r.get('address') == ip and r.get('list') == 'bloqueados':
                lista.remove(id=r['.id'])
                print(f"[OK] IP {ip} eliminada de lista 'bloqueados'")
            else:
                print(f"[IGNORADO] No coincide: {r.get('address')} en lista {r.get('list')}")

        pool.disconnect()
    except Exception as e:
        print(f"[ERROR] Al reconectar {ip}: {str(e)}")
    return redirect(url_for('lista_clientes'))

@app.route('/neighbors')
def neighbors():
    neighbors_data = mikrotik_execute('/ip/neighbor', 'list')
    print(neighbors_data)  # Para revisar en consola
    return render_template('neighbors.html', neighbors=neighbors_data)


# Ruta 1: página HTML visible
@app.route('/consumo')
def ver_consumo():
    return render_template('consumo.html')  # Solo carga la plantilla HTML


def formatear_velocidad(rate_str):
    try:
        if not rate_str or '/' not in rate_str:
            return "Sin tráfico"
        
        up_str, down_str = rate_str.strip().split('/')

        # Si por alguna razón vienen vacíos
        up = int(up_str) if up_str.strip().isdigit() else 0
        down = int(down_str) if down_str.strip().isdigit() else 0

        def formato(bits):
            if bits >= 1_000_000:
                return f"{round(bits / 1_000_000, 2)} Mbps"
            elif bits >= 1_000:
                return f"{round(bits / 1_000, 2)} Kbps"
            else:
                return f"{bits} bps"

        return f"{formato(up)} / {formato(down)}"
    except Exception as e:
        print("Error formateando velocidad:", e)
        return "Sin tráfico"

@app.route('/api/consumo', methods=['GET'])
def api_consumo():
    consumo = []

    try:
        pool, api = conectar_api()
        sq = api.get_resource('/queue/simple')
        queues = sq.get()

        for q in queues:
            name = q.get('name', '---')
            target = q.get('target', '---')
            rate_str = q.get('rate', '0/0')  # Normalmente viene "upload/download"
            bytes_str = q.get('bytes', '0/0')

            # Parse upload/download de bytes
            upload_bytes = 0
            download_bytes = 0
            if '/' in bytes_str:
                partes = bytes_str.split('/')
                try:
                    upload_bytes = int(partes[0])
                    download_bytes = int(partes[1])
                except:
                    pass

            # Formatear velocidad para mostrar
            def formato(bits):
                if bits >= 1_000_000:
                    return f"{round(bits / 1_000_000, 2)} Mbps"
                elif bits >= 1_000:
                    return f"{round(bits / 1_000, 2)} Kbps"
                else:
                    return f"{bits} bps"

            if '/' in rate_str:
                try:
                    up_rate_str, down_rate_str = rate_str.split('/')
                    up_rate = int(up_rate_str)
                    down_rate = int(down_rate_str)
                    rate = f"{formato(up_rate)} / {formato(down_rate)}"
                except:
                    rate = "0 bps / 0 bps"
            else:
                rate = "0 bps / 0 bps"

            consumo.append({
                'name': name,
                'target': target,
                'upload': round(upload_bytes / (1024 * 1024), 2),    # MB
                'download': round(download_bytes / (1024 * 1024), 2), # MB
                'bytes': round((upload_bytes + download_bytes) / (1024 * 1024), 2),  # Total MB
                'rate': rate,
                '.id': q.get('.id')
            })

        pool.disconnect()
        return jsonify({'error': None, 'data': consumo})

    except Exception as e:
        mensaje = str(e)

        if 'invalid user name or password' in mensaje:
            mensaje_amigable = 'Credenciales incorrectas. Verifica usuario y contraseña del MikroTik.'
        elif 'timed out' in mensaje or 'TimeoutError' in mensaje:
            mensaje_amigable = 'Tiempo de espera agotado. No se pudo conectar con el MikroTik.'
        elif 'Connection refused' in mensaje:
            mensaje_amigable = 'Conexión rechazada. Revisa si el MikroTik está encendido y accesible.'
        else:
            mensaje_amigable = f'Error al obtener datos del MikroTik.'

        return jsonify({'error': mensaje_amigable, 'data': []})

    
@app.route('/queues')
def queues_page():
    return render_template('queues.html')

def formatear_velocidad(valor):
    mapa = {
        '0': '0',
        '1000000': '1M',
        '2000000': '2M',
        '5000000': '5M',
        '10000000': '10M',
        '15000000': '15M',
        '20000000': '20M',
        '25000000': '25M',
        '30000000': '30M'
    }
    return mapa.get(valor, '0')  # Si no coincide, devuelve '0' como sin límite

def ping_ip(ip):
    try:
        count_param = '-n' if platform.system().lower() == 'windows' else '-c'
        result = subprocess.run(['ping', count_param, '4', ip], capture_output=True, text=True, timeout=10)
        output = result.stdout

        avg_time = None
        has_less_than_1ms = False
        packets_received = None

        lines = output.splitlines()

        if platform.system().lower() == 'windows':
            # Buscar paquetes recibidos
            for line in lines:
                if "recibidos" in line.lower():
                    match = re.search(r"recibidos\s*=\s*(\d+)", line.lower())
                    if match:
                        packets_received = int(match.group(1))

            # Buscar ping promedio y <1ms
            for line in lines:
                if "tiempo<1m" in line.lower():
                    has_less_than_1ms = True
                if "media" in line.lower() or "average" in line.lower():
                    match = re.search(r"(media|average)\s*=\s*(\d+)", line.lower())
                    if match:
                        avg_time = int(match.group(2))
                        break

            if packets_received == 0:
                return {'success': False, 'error': 'Host no responde (0 paquetes recibidos)', 'raw_output': output}

            if avg_time == 0 and has_less_than_1ms:
                avg_time = "<1"

        else:
            # Linux
            for line in lines:
                if 'rtt min/avg/max/mdev' in line:
                    stats = line.split('=')[1].split('/')
                    avg_time = stats[1].strip()
                    break

            for line in lines:
                if "packet loss" in line:
                    match = re.search(r"(\d+)% packet loss", line)
                    if match and int(match.group(1)) == 100:
                        return {'success': False, 'error': 'Host no responde (100% pérdida de paquetes)', 'raw_output': output}

        return {'success': True, 'avg_ping_ms': avg_time, 'raw_output': output}

    except subprocess.TimeoutExpired:
        return {'success': False, 'error': 'El ping ha excedido el tiempo de espera (timeout). Verifique que la IP esté activa o accesible.'}
    except FileNotFoundError:
        return {'success': False, 'error': 'El comando ping no se encontró en el sistema.'}
    except Exception as e:
        msg = str(e)
        if 'timed out' in msg.lower():
            return {'success': False, 'error': 'Tiempo de espera agotado al intentar hacer ping. Por favor, verifique la IP.'}
        return {'success': False, 'error': f'Ocurrió un error inesperado al hacer ping: {msg}'}


def run_speedtest():
    try:
        st = speedtest.Speedtest()
        st.get_best_server()
        download_speed = st.download()
        upload_speed = st.upload()
        return {
            'success': True,
            'download_mbps': round(download_speed / 1_000_000, 2),
            'upload_mbps': round(upload_speed / 1_000_000, 2)
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/test', methods=['GET', 'POST'])
def test_speed():
    if request.method == 'GET':
        return render_template('test_form.html')

    ip = request.form.get('ip')

    ping_result = ping_ip(ip)
    speed_result = run_speedtest()

    if not ping_result['success']:
        return render_template('test_form.html', result={'error': f"Error de ping: {ping_result['error']}"}, ip=ip)

    if not speed_result['success']:
        return render_template('test_form.html', result={'error': f"Error de speedtest: {speed_result['error']}"}, ip=ip)

    return render_template(
        'test_form.html',
        result={
            'ping_ms': ping_result['avg_ping_ms'],
            'raw_ping': ping_result['raw_output'],
            'download_mbps': speed_result['download_mbps'],
            'upload_mbps': speed_result['upload_mbps']
        },
        ip=ip
    )


@app.route('/api/queues')
def obtener_queues():
    try:
        pool, api = conectar_api()
        resource = api.get_resource('/queue/simple')
        queues_data = resource.get()

        # Obtener lista de IPs cortadas (address-list)
        cortados_resource = api.get_resource('/ip/firewall/address-list')
        cortados = cortados_resource.get(list='bloqueados')  # Asumiendo lista 'bloqueados'
        ips_cortadas = {c['address'] for c in cortados}

        queues = []
        for q in queues_data:
            upload, download = ('0', '0')
            if 'max-limit' in q and '/' in q['max-limit']:
              raw_up, raw_down = q['max-limit'].split('/')
              upload = raw_up.strip()
              download = raw_down.strip()

            # Extraemos solo la IP sin máscara para comparación y para mostrar
            target_ip = q.get('target', '').split('/')[0]

            estado = "Activo"
            if q.get('disabled', 'false') == 'true':
                estado = "Inactivo"
            elif target_ip in ips_cortadas:
                estado = "Cortado"

            queues.append({
                'id': q.get('.id') or q.get('id') or '',
                'name': q.get('name', ''),
                'target': target_ip,  # Guardamos sin máscara aquí
                'upload_limit': upload,
                'download_limit': download,
                'estado': estado
            })

        pool.disconnect()
        return jsonify({'data': queues})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/queues/cortar', methods=['POST'])
def cortar_queue():
    data = request.get_json()
    ip = data.get('ip')
    if not ip:
        return jsonify({'error': 'Falta IP'}), 400
    try:
        pool, api = conectar_api()
        addr_list = api.get_resource('/ip/firewall/address-list')
        # Agregar IP a la lista "bloqueados"
        addr_list.add(address=ip, list='bloqueados', comment='Corte automático')
        pool.disconnect()
        return jsonify({'status': 'ok'})
    except Exception as e:
        print(f"[ERROR cortar_queue] {e}")  # 👈 Esto mostrará el error real en consola
        return jsonify({'error': str(e)}), 500


@app.route('/api/queues/reconectar', methods=['POST'])
def reconectar_queue():
    data = request.get_json()
    ip = data.get('ip')
    
    if not ip:
        return jsonify({'error': 'Falta IP'}), 400

    try:
        pool, api = conectar_api()
        addr_list = api.get_resource('/ip/firewall/address-list')

        registros = addr_list.get()
        encontrada = False
        for r in registros:
            if r.get('address') == ip and r.get('list') == 'bloqueados':
                addr_list.remove(id=r['id'])
                encontrada = True
                break

        pool.disconnect()

        if encontrada:
            return jsonify({'status': 'ok', 'mensaje': 'IP eliminada de la lista bloqueados'})
        else:
            return jsonify({'status': 'ok', 'mensaje': 'La IP no estaba en la lista'})

    except Exception as e:
        print(f"[ERROR reconectar_queue] {e}")  # 👈 Este print te mostrará el error real en consola
        return jsonify({'error': str(e)}), 500


@app.route('/api/queues/agregar', methods=['POST'])
def agregar_queue():
    data = request.get_json()
    try:
        name = data['name']
        target = data['target']
        upload = data['upload_limit']
        download = data['download_limit']
        max_limit = f"{upload}/{download}"

        pool, api = conectar_api()
        q = api.get_resource('/queue/simple')
        q.add(name=name, target=target, **{'max-limit': max_limit})
        pool.disconnect()

        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/queues/editar', methods=['POST'])
def editar_queue():
    data = request.get_json()
    queue_id = data.get('id')
    upload_limit = data.get('upload_limit')
    download_limit = data.get('download_limit')

    if not queue_id or not upload_limit or not download_limit:
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        pool, api = conectar_api()
        resource = api.get_resource('/queue/simple')

        # Verificar que exista
        existing = resource.get(id=queue_id)
        if not existing:
            pool.disconnect()
            return jsonify({'error': 'ID de queue no encontrado'}), 404

        max_limit_str = f"{upload_limit}/{download_limit}"
        resource.set(id=queue_id, **{'max-limit': max_limit_str})

        pool.disconnect()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/queues/eliminar', methods=['POST'])
def eliminar_queue():
    data = request.get_json()
    queue_id = data.get('id')

    if not queue_id:
        return jsonify({'error': 'Falta el id'}), 400

    try:
        pool, api = conectar_api()
        resource = api.get_resource('/queue/simple')

        # Verificar que exista
        existing = resource.get(id=queue_id)
        if not existing:
            pool.disconnect()
            return jsonify({'error': 'ID de queue no encontrado'}), 404

        resource.remove(id=queue_id)
        pool.disconnect()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Configuración de MikroTik
MikroTik_IP = '192.168.88.1'
MikroTik_User = 'Francisco'
MikroTik_Pass = '1251301881'
MikroTik_Port = 8728  # Puerto por defecto de la API

# Función para ejecutar comandos en MikroTik
def mikrotik_execute(resource_name, action, params=None):
    try:
        connection = RouterOsApiPool(
            host=MikroTik_IP,
            username=MikroTik_User,
            password=MikroTik_Pass,
            port=MikroTik_Port,
            plaintext_login=True
        )
        api = connection.get_api()
        resource = api.get_resource(resource_name)

        if action == "list":
            result = resource.get()
        elif action == "add" and params:
            result = resource.add(**params)
        elif action == "remove" and params:
            resource.remove(id=params['id'])
            result = {"status": "success"}
        elif action == "set" and params:
            resource.set(**params)
            result = {"status": "success"}
        elif action == "reboot":
            system_resource = api.get_resource("/system/reboot")
            result = system_resource.call("execute", {})
        elif action == "reset":
            system_resource = api.get_resource("/system/reset-configuration")
            result = system_resource.call("execute", {})
        else:
            result = {"error": "Acción no soportada."}

        connection.disconnect()
        return result
    except Exception as e:
        return {"error": str(e)}

# Función para obtener el log de MikroTik
def get_mikrotik_logs():
    try:
        connection = RouterOsApiPool(
            host=MikroTik_IP,
            username=MikroTik_User,
            password=MikroTik_Pass,
            port=MikroTik_Port,
            plaintext_login=True
        )
        api = connection.get_api()
        log_resource = api.get_resource("/log")
        logs = log_resource.get()
        connection.disconnect()
        return logs
    except Exception as e:
        return {"error": str(e)}

# Ruta para la página del log
@app.route('/log')
def log():
    return render_template('log.html')

# Ruta para obtener los logs como JSON
@app.route('/api/logs', methods=['GET'])
def logs_api():
    logs = get_mikrotik_logs()
    return jsonify(logs)

# Página principal: Lista los usuarios del Hotspot
@app.route('/hotspot')
def hotspot():
    # Obtener la lista de usuarios del hotspot
    users = mikrotik_execute('/ip/hotspot/user', 'list')

    # Obtener la lista de perfiles para mapear el rate-limit
    profiles = mikrotik_execute('/ip/hotspot/user/profile', 'list')
    profile_map = {profile['name']: profile.get('rate-limit', 'No Configurado') for profile in profiles}

    # Obtener las sesiones activas para cada usuario
    active_sessions = mikrotik_execute('/ip/hotspot/active', 'list')
    session_map = {session['user']: session.get('session-time-left', 'No Activo') for session in active_sessions}

    # Enriquecer los usuarios con la información de perfil y sesiones activas
    for user in users:
        profile_name = user.get('profile', 'Default')
        user['rate-limit'] = profile_map.get(profile_name, 'No Configurado')
        user['session-time-left'] = session_map.get(user['name'], 'No Activo')

    if 'error' in users:
        return f"Error al obtener usuarios: {users['error']}"

    return render_template('hotspot.html', users=users)

@app.route('/active_sessions')
def active_sessions():
    # Obtener las sesiones activas
    active_sessions = mikrotik_execute('/ip/hotspot/active', 'list')
    
    if 'error' in active_sessions:
        return {"error": active_sessions['error']}, 500

    # Crear un diccionario con el tiempo restante por usuario
    session_data = {
        session['user']: session.get('session-time-left', 'No Activo')
        for session in active_sessions
    }
    return session_data

@app.route('/hotspot_monitoring')
def hotspot_monitoring():
    # Obtener dispositivos activos en el Hotspot
    active_hosts = mikrotik_execute('/ip/hotspot/active', 'list')
    if 'error' in active_hosts:
        return f"Error al obtener la información del Hotspot: {active_hosts['error']}"

    # Obtener la información del DHCP Server Lease
    dhcp_leases = mikrotik_execute('/ip/dhcp-server/lease', 'list')
    if 'error' in dhcp_leases:
        return f"Error al obtener la información de DHCP Server: {dhcp_leases['error']}"

    # Vincular el hostname y determinar el estado basado en el DHCP Lease
    for host in active_hosts:
        # Buscar la arrendación que coincide con la dirección IP
        matching_lease = next((lease for lease in dhcp_leases if lease.get('address') == host.get('address')), None)
        if matching_lease:
            # Asignar el hostname desde DHCP Lease si no existe
            host['host-name'] = matching_lease.get('host-name', 'Hostname no disponible')
            # Asignar el estado basado en si está vinculado ("bound")
            host['status'] = "online" if matching_lease.get('status') == "bound" else "offline"
        else:
            # Si no hay arrendación DHCP correspondiente, marcar como offline
            host['status'] = "offline"

    return render_template('hotspot_monitoring.html', hosts=active_hosts)


# Crear un nuevo usuario
@app.route('/add_user', methods=['POST'])
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    time_limit = request.form.get('time_limit')
    rate_limit = request.form.get('rate_limit')  # Control de velocidad

    # Verificar si el perfil ya existe, si no, crearlo
    profile_name = f"profile_{rate_limit.replace('/', '_')}"  # Ejemplo: profile_2M_2M
    existing_profiles = mikrotik_execute('/ip/hotspot/user/profile', 'list')

    if not any(profile['name'] == profile_name for profile in existing_profiles):
        # Crear el perfil con el límite de velocidad
        profile_params = {
            'name': profile_name,
            'rate-limit': rate_limit
        }
        profile_result = mikrotik_execute('/ip/hotspot/user/profile', 'add', profile_params)
        if 'error' in profile_result:
            return f"Error al crear el perfil: {profile_result['error']}"

    # Crear el usuario y asignar el perfil
    params = {
        'name': username,
        'password': password,
        'limit-uptime': time_limit,
        'profile': profile_name  # Asignar el perfil al usuario
    }

    result = mikrotik_execute('/ip/hotspot/user', 'add', params)
    if 'error' in result:
        return f"Error al agregar el usuario: {result['error']}"
    return redirect(url_for('hotspot'))


# Eliminar un usuario
@app.route('/delete_user/<user_id>', methods=['POST'])
def delete_user(user_id):
    params = {'id': user_id}
    result = mikrotik_execute('/ip/hotspot/user', 'remove', params)
    if 'error' in result:
        return f"Error al eliminar el usuario: {result['error']}"
    return redirect(url_for('hotspot'))

import paramiko

@app.route('/terminal', methods=['GET'])
def terminal():
    return render_template('terminal.html')

@app.route('/api/terminal/run', methods=['POST'])
def run_terminal_command():
    data = request.get_json()
    command = data.get('command', '').strip()
    output = ''
    
    if not command:
        return jsonify({'output': 'Ingresa un comando válido.'})
    
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname='192.168.88.1',
            username='Francisco',
            password='1251301881',
            port=22  # ← Usa el puerto 22 aquí
        )

        stdin, stdout, stderr = ssh.exec_command(command)
        output = stdout.read().decode() + stderr.read().decode()
        ssh.close()

        # Guardar historial en sesión (últimos 20)
        history = session.get('history', [])
        history.append(command)
        session['history'] = history[-20:]

    except Exception as e:
        output = f"Error: {str(e)}"

    return jsonify({'output': output})

import re
from flask import jsonify

# Limpia el ID MikroTik para que solo quede el número o identificador sin asterisco ni espacios
def limpiar_id(rule_id):
    if not rule_id:
        return ""
    return rule_id.lstrip('*').strip()

@app.route('/firewall', methods=['GET'])
def ver_firewall():
    rules = []
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname='192.168.88.1',
            username='Francisco',
            password='1251301881',
            port=22
        )

        stdin, stdout, stderr = ssh.exec_command('/ip firewall filter print detail without-paging')
        output = stdout.read().decode()
        ssh.close()

        current_rule = {}

        for line in output.splitlines():
            line = line.strip()
            if not line or line.startswith("Flags"):
                continue

            # Detectar nueva regla: línea que comienza con número
            if line and line[0].isdigit():
                # Guardar regla previa
                if current_rule:
                    rules.append(current_rule)
                    current_rule = {}

                parts = line.split(' ', 2)
                raw_id = parts[0]
                current_rule["id"] = limpiar_id(raw_id)
                rest_line = parts[2] if len(parts) > 2 else ""

                if ';;;' in rest_line:
                    props_part, comment_part = rest_line.split(';;;', 1)
                    current_rule["comment"] = comment_part.strip()
                else:
                    props_part = rest_line

                for prop in props_part.strip().split():
                    if '=' in prop:
                        key, value = prop.split('=', 1)
                        current_rule[key] = value
            else:
                if ';;;' in line:
                    comment_part = line.split(';;;', 1)[1].strip()
                    if "comment" in current_rule:
                        current_rule["comment"] += " " + comment_part
                    else:
                        current_rule["comment"] = comment_part
                else:
                    for prop in line.split():
                        if '=' in prop:
                            key, value = prop.split('=', 1)
                            current_rule[key] = value

        if current_rule:
            rules.append(current_rule)

    except Exception as e:
        rules = [{'error': str(e)}]

    print("Reglas cargadas:", rules)
    return render_template('firewall.html', parsed_rules=rules)


def ejecutar_comando_mikrotik(comando):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname='192.168.88.1', username='Francisco', password='1251301881', port=22)
    stdin, stdout, stderr = ssh.exec_command(comando)
    output = stdout.read().decode()
    ssh.close()
    return output

@app.route('/api/firewall/create', methods=['POST'])
def crear_regla():
    data = request.json
    try:
        cmd = '/ip firewall filter add '
        cmd += f"chain={data.get('chain')} action={data.get('action')} "

        if data.get('protocol'):
            cmd += f"protocol={data.get('protocol')} "
        if data.get('dst-port'):
            cmd += f"dst-port={data.get('dst-port')} "
            if data.get('comment'):
            # Los comentarios se ponen con ;;; en MikroTik, ejemplo:
              cmd += f' comment="{data.get("comment")}" '

        ejecutar_comando_mikrotik(cmd.strip())
        return jsonify(success=True, message="Regla creada exitosamente")
    except Exception as e:
        return jsonify(success=False, message=str(e))


@app.route('/api/firewall/enable', methods=['POST'])
def activar_regla():
    data = request.json
    try:
        rule_id = limpiar_id(data.get('id'))
        if not rule_id:
            return jsonify(success=False, message="ID de regla es requerido")

        cmd = f"/ip firewall filter enable {rule_id}"
        ejecutar_comando_mikrotik(cmd)

        return jsonify(success=True, message="Regla activada")
    except Exception as e:
        return jsonify(success=False, message=str(e))


@app.route('/api/firewall/update', methods=['POST'])
def actualizar_regla():
    data = request.json
    try:
        rule_id = data.get('id')
        if not rule_id:
            return jsonify(success=False, message="ID de regla es requerido")

        cmd = f"/ip firewall filter set {rule_id} "

        if 'chain' in data and data['chain']:
            cmd += f"chain={data['chain']} "
        if 'action' in data and data['action']:
            cmd += f"action={data['action']} "
        if 'protocol' in data and data['protocol']:
            cmd += f"protocol={data['protocol']} "
        if 'dst-port' in data and data['dst-port']:
            cmd += f"dst-port={data['dst-port']} "

        if 'comment' in data:
            if data['comment']:
                cmd += f'comment="{data["comment"]}" '
            else:
                cmd += 'comment="" '

        print("Comando a ejecutar:", cmd.strip())
        ejecutar_comando_mikrotik(cmd.strip())
        return jsonify(success=True, message="Regla actualizada exitosamente")
    except Exception as e:
        return jsonify(success=False, message=str(e))


@app.route('/api/firewall/disable', methods=['POST'])
def desactivar_regla():
    data = request.json
    try:
        rule_id = limpiar_id(data.get('id'))
        if not rule_id:
            return jsonify(success=False, message="ID de regla es requerido")

        cmd = f"/ip firewall filter disable {rule_id}"
        ejecutar_comando_mikrotik(cmd)

        return jsonify(success=True, message="Regla desactivada")
    except Exception as e:
        return jsonify(success=False, message=str(e))


# Nuevo endpoint para eliminar regla
@app.route('/api/firewall/delete', methods=['POST'])
def eliminar_regla():
    data = request.json
    try:
        rule_id = limpiar_id(data.get('id'))
        if not rule_id:
            return jsonify(success=False, message="ID de regla es requerido")

        cmd = f"/ip firewall filter remove {rule_id}"
        ejecutar_comando_mikrotik(cmd)

        return jsonify(success=True, message="Regla eliminada exitosamente")
    except Exception as e:
        return jsonify(success=False, message=str(e))

@app.route('/usuarios')
def usuarios_page():
    return render_template('usuarios_list.html')

# Página lista usuarios (admin)
@app.route('/api/usuarios', methods=['GET'])
def listar_usuarios():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, correo, telefono, rol, area, username FROM usuarios ORDER BY id DESC")
        filas = cur.fetchall()
        cur.close()
        conn.close()

        usuarios = []
        for fila in filas:
            usuarios.append({
                "id": fila[0],
                "nombre": fila[1],
                "correo": fila[2],
                "telefono": fila[3],
                "rol": fila[4],
                "area": fila[5],
                "username": fila[6]
            })

        return jsonify(usuarios)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def generar_username(nombre):
    base = ''.join(e for e in nombre.lower() if e.isalnum())
    sufijo = ''.join(random.choices(string.digits, k=3))
    return f"{base}{sufijo}"

def generar_password(length=10):
    chars = string.ascii_letters + string.digits + "!@#$%&*"
    return ''.join(random.choices(chars, k=length))

@app.route('/api/usuarios', methods=['POST'])
def crear_usuario():
    data = request.json
    nombre = data.get('nombre')
    correo = data.get('correo')
    telefono = data.get('telefono')
    rol = data.get('rol')
    area = data.get('area')

    if not nombre or not correo:
        return jsonify({"error": "Nombre y correo son obligatorios"}), 400

    username = generar_username(nombre)
    password = generar_password()
    password_hash = generate_password_hash(password)

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO usuarios (nombre, correo, telefono, rol, area, username, password_hash)
            VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (nombre, correo, telefono, rol, area, username, password_hash))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({
            "id": new_id,
            "nombre": nombre,
            "correo": correo,
            "telefono": telefono,
            "rol": rol,
            "area": area,
            "username": username,
            "password": password  # se devuelve password en texto claro SOLO AQUÍ
        }), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# PUT y DELETE para editar y eliminar usuarios no cambian credenciales
@app.route('/api/usuarios/<int:id>', methods=['PUT'])
def editar_usuario(id):
    data = request.json
    nombre = data.get('nombre')
    correo = data.get('correo')
    telefono = data.get('telefono')
    rol = data.get('rol')
    area = data.get('area')
    if not nombre or not correo:
        return jsonify({"error": "Nombre y correo son obligatorios"}), 400
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE usuarios SET nombre=%s, correo=%s, telefono=%s, rol=%s, area=%s WHERE id=%s
        """, (nombre, correo, telefono, rol, area, id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Usuario actualizado correctamente"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/usuarios/<int:id>', methods=['DELETE'])
def eliminar_usuario(id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM usuarios WHERE id=%s", (id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Usuario eliminado correctamente"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/usuarios/<int:id>', methods=['GET'])
def obtener_usuario(id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, correo, telefono, rol, area, username FROM usuarios WHERE id=%s", (id,))
        fila = cur.fetchone()
        cur.close()
        conn.close()
        if fila:
            usuario = {
                "id": fila[0],
                "nombre": fila[1],
                "correo": fila[2],
                "telefono": fila[3],
                "rol": fila[4],
                "area": fila[5],
                "username": fila[6]
            }
            return jsonify(usuario)
        else:
            return jsonify({"error": "Usuario no encontrado"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/usuarios/<int:id>/reset_password', methods=['POST'])
def resetear_password(id):
    try:
        nueva_password = generar_password()
        password_hash = generate_password_hash(nueva_password)

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET password_hash = %s WHERE id = %s", (password_hash, id))
        conn.commit()
        cur.close()
        conn.close()

        # Devuelve la nueva contraseña en texto plano SOLO AQUÍ
        return jsonify({"contrasena": nueva_password})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

from werkzeug.security import check_password_hash
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, username, password_hash, rol, area FROM usuarios WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[3], password):
            session['usuario_id'] = user[0]
            session['nombre'] = user[1]
            session['username'] = user[2]
            session['rol'] = user[4]
            session['area'] = user[5]
            return redirect(url_for('index'))  # o ruta principal tras login
        else:
            flash('Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


def rol_requerido(*roles_permitidos):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('rol') not in roles_permitidos:
                return "Acceso denegado", 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def area_requerida(*areas_permitidas):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('area') not in areas_permitidas:
                return "Acceso denegado", 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/inventario')
def ver_inventario():
    codigo = request.args.get('codigo')
    categoria = request.args.get('categoria')
    estado = request.args.get('estado')
    ubicacion = request.args.get('ubicacion')
    proveedor = request.args.get('proveedor')

    query = "SELECT * FROM inventario WHERE TRUE"
    params = []

    if codigo:
        query += " AND codigo = %s"
        params.append(codigo)
    if categoria:
        query += " AND categoria = %s"
        params.append(categoria)
    if estado:
        query += " AND estado = %s"
        params.append(estado)
    if ubicacion:
        query += " AND ubicacion = %s"
        params.append(ubicacion)
    if proveedor:
        query += " AND proveedor = %s"
        params.append(proveedor)

    query += " ORDER BY fecha_ingreso DESC"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, params)
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    inventario = [dict(zip(columnas, fila)) for fila in filas]

    # Obtener series y MACs para los productos filtrados
    producto_ids = [item['id'] for item in inventario]
    detalles_por_producto = {}

    if producto_ids:
        # Para evitar problemas con listas vacías
        format_strings = ','.join(['%s'] * len(producto_ids))
        cur.execute(
            f"SELECT producto_id, numero_serie, mac_address FROM detalles_producto WHERE producto_id IN ({format_strings})",
            tuple(producto_ids)
        )
        detalles_raw = cur.fetchall()
        for producto_id, serie, mac in detalles_raw:
            detalles_por_producto.setdefault(producto_id, []).append({'serie': serie, 'mac': mac})

    # Agregar detalles a cada producto
    for item in inventario:
        item['detalles'] = detalles_por_producto.get(item['id'], [])
        # Para que item.detalles esté disponible y se llame item.series (alias)
    for item in inventario:
        item['series'] = item.get('detalles', [])

    cur.close()
    conn.close()

    return render_template('inventario.html', inventario=inventario)

@app.route('/inventario/agregar', methods=['POST'])
def agregar_producto():
    imagen = request.files.get('imagen_domicilio')
    imagen_nombre = None
    datos = request.form
    codigo = generar_codigo_unico()

    series = request.form.getlist('series[]')
    macs = request.form.getlist('macs[]')
    errores = []

    print("Series recibidas:", series)
    print("MACs recibidas:", macs)

    if imagen and imagen.filename != '':
        filename = secure_filename(imagen.filename)
        imagen_nombre = filename
        imagen.save(os.path.join('static/uploads', filename))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO inventario (
                codigo, nombre, descripcion, categoria, cantidad, unidad, marca, modelo,
                numero_serie, mac_address, especificaciones, ubicacion, proveedor, estado, imagen_domicilio, precio_unitario
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, '', %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            codigo,
            datos.get('nombre'),
            datos.get('descripcion'),
            datos.get('categoria'),
            len(series),
            datos.get('unidad'),
            datos.get('marca'),
            datos.get('modelo'),
            '',  # numero_serie en inventario
            '',  # mac_address en inventario
            datos.get('ubicacion'),
            datos.get('proveedor'),
            datos.get('estado'),
            imagen_nombre,
            datos.get('precio_unitario')
        ))
        producto_id = cur.fetchone()[0]

        for serie, mac in zip(series, macs):
            serie = serie.strip()
            mac = mac.strip()
            print(f"Insertando serie {serie} y mac {mac}")

            cur.execute("SELECT 1 FROM detalles_producto WHERE mac_address = %s", (mac,))
            if cur.fetchone():
                errores.append(f"La MAC {mac} ya está registrada.")
            else:
                cur.execute("""
                    INSERT INTO detalles_producto (producto_id, numero_serie, mac_address)
                    VALUES (%s, %s, %s)
                """, (producto_id, serie, mac))

        conn.commit()

        if errores:
            flash("Errores: " + "; ".join(errores), 'danger')
        else:
            flash("Producto registrado exitosamente con sus series/MACs", 'success')

    except Exception as e:
        conn.rollback()
        print("Error:", e)
        flash(f"Error al registrar producto: {e}", 'danger')

    finally:
        cur.close()
        conn.close()

    return redirect(url_for('ver_inventario'))


@app.route('/inventario/editar/<int:id>', methods=['POST'])
def editar_producto(id):
    datos = request.form
    imagen = request.files.get('imagen_domicilio')
    imagen_nombre = None

    # Procesar imagen si se sube una nueva
    if imagen and imagen.filename != '':
        filename = secure_filename(imagen.filename)
        imagen_nombre = filename
        imagen.save(os.path.join('static/uploads', filename))

    series = request.form.getlist('series[]')
    macs = request.form.getlist('macs[]')

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # Actualizar producto principal
        cur.execute("""
            UPDATE inventario SET nombre=%s, descripcion=%s, categoria=%s, cantidad=%s, unidad=%s,
                marca=%s, modelo=%s, numero_serie=%s, mac_address=%s, ubicacion=%s,
                proveedor=%s, estado=%s {imagen_sql}, precio_unitario=%s
            WHERE id=%s
        """.format(
            imagen_sql=", imagen_domicilio=%s" if imagen_nombre else ""
        ), tuple([
            datos.get('nombre'),
            datos.get('descripcion'),
            datos.get('categoria'),
            int(datos.get('cantidad')),
            datos.get('unidad'),
            datos.get('marca'),
            datos.get('modelo'),
            '',  # Vacío porque usamos detalles_producto
            '',
            datos.get('ubicacion'),
            datos.get('proveedor'),
            datos.get('estado'),
            imagen_nombre,
            datos.get('precio_unitario'),
            id
        ] if imagen_nombre else [
            datos.get('nombre'),
            datos.get('descripcion'),
            datos.get('categoria'),
            int(datos.get('cantidad')),
            datos.get('unidad'),
            datos.get('marca'),
            datos.get('modelo'),
            '',
            '',
            datos.get('ubicacion'),
            datos.get('proveedor'),
            datos.get('estado'),
            datos.get('precio_unitario'),
            id
        ]))

        # Eliminar series/mac anteriores
        cur.execute("DELETE FROM detalles_producto WHERE producto_id = %s", (id,))

        # Insertar nuevas series/mac
        for serie, mac in zip(series, macs):
            cur.execute("""
                INSERT INTO detalles_producto (producto_id, numero_serie, mac_address)
                VALUES (%s, %s, %s)
            """, (id, serie.strip(), mac.strip()))

        conn.commit()
        flash("Producto actualizado correctamente.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error al actualizar: {e}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for('ver_inventario'))

@app.route('/inventario/eliminar/<int:id>', methods=['POST'])
def eliminar_producto(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM inventario WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('ver_inventario'))

import random
def generar_codigo_unico():
    conn = get_db_connection()
    cur = conn.cursor()
    while True:
        codigo = f"{random.randint(0, 9999):04d}"  # 4 dígitos con ceros a la izquierda
        cur.execute("SELECT 1 FROM inventario WHERE codigo = %s", (codigo,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return codigo
    # Nunca debería llegar aquí

from openpyxl import Workbook
from fpdf import FPDF
import os, io
from flask import send_file
@app.route('/exportar_excel')
def exportar_excel():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT codigo, nombre, categoria, cantidad, unidad, marca, modelo, ubicacion, proveedor, estado FROM inventario ORDER BY fecha_ingreso DESC")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Header
    headers = ['Código', 'Nombre', 'Categoría', 'Cantidad', 'Unidad', 'Marca', 'Modelo', 'Ubicación', 'Proveedor', 'Estado']
    ws.append(headers)

    for row in rows:
        ws.append(row)

    # Guardar archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="inventario.xlsx", as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/exportar_pdf')
def exportar_pdf():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT codigo, nombre, categoria, cantidad, unidad, marca, modelo, ubicacion, proveedor, estado, imagen_domicilio
        FROM inventario ORDER BY fecha_ingreso DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Inventario", ln=True, align='C')
    pdf.ln(10)

    for row in rows:
        codigo, nombre, categoria, cantidad, unidad, marca, modelo, ubicacion, proveedor, estado, imagen = row

        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, f"{nombre} ({codigo})", ln=True)

        # Mostrar imagen si existe
        if imagen:
            img_path = os.path.join('static', 'uploads', imagen)
            if os.path.isfile(img_path):
                # Ajusta tamaño y posición si quieres
                pdf.image(img_path, w=40)
        
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 8, f"Categoría: {categoria}", ln=True)
        pdf.cell(0, 8, f"Marca/Modelo: {marca} {modelo}", ln=True)
        pdf.cell(0, 8, f"Cantidad: {cantidad} {unidad}", ln=True)
        pdf.cell(0, 8, f"Ubicación: {ubicacion}", ln=True)
        pdf.cell(0, 8, f"Proveedor: {proveedor}", ln=True)
        pdf.cell(0, 8, f"Estado: {estado}", ln=True)
        pdf.ln(10)

    # Exportar PDF a bytes y crear un BytesIO
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output = io.BytesIO(pdf_bytes)
    pdf_output.seek(0)

    return send_file(pdf_output, download_name="inventario.pdf", as_attachment=True, mimetype="application/pdf")

@app.route('/facturas/nueva', methods=['GET', 'POST'])
def nueva_factura():
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener contratos activos para seleccionar cliente
    cur.execute("SELECT id, numero_contrato, nombres, apellidos FROM contratos WHERE estado = 'Activo'")
    contratos = cur.fetchall()

    # Obtener productos disponibles para añadir a la factura
    cur.execute("SELECT id, nombre, marca, modelo, precio_unitario FROM inventario WHERE cantidad > 0")
    productos = cur.fetchall()

    if request.method == 'POST':
        datos = request.form
        contrato_id = datos.get('contrato_id')
        numero_factura = datos.get('numero_factura').strip()
        notas = datos.get('notas', '')

        # Recibir listas dinámicas desde el formulario
        productos_ids = request.form.getlist('producto_id[]')
        cantidades = request.form.getlist('cantidad[]')

        subtotal = 0
        detalles = []

        for pid, cant_str in zip(productos_ids, cantidades):
            try:
                cantidad = int(cant_str)
            except ValueError:
                cantidad = 0

            # Obtener nombre y precio unitario
            cur.execute("SELECT nombre, precio_unitario FROM inventario WHERE id = %s", (pid,))
            row = cur.fetchone()
            if row:
                nombre_producto, precio_unitario = row
                precio_unitario = float(precio_unitario)
                item_subtotal = round(precio_unitario * cantidad, 2)
                subtotal += item_subtotal
                detalles.append((pid, nombre_producto, cantidad, precio_unitario, item_subtotal))

        subtotal = round(subtotal, 2)
        impuestos = round(subtotal * 0.15, 2)  # IVA del 15%
        total = round(subtotal + impuestos, 2)

        try:
            # Insertar la factura
            cur.execute("""
                INSERT INTO facturas (contrato_id, numero_factura, subtotal, impuestos, total, estado, notas)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (contrato_id, numero_factura, subtotal, impuestos, total, 'Pendiente', notas))
            factura_id = cur.fetchone()[0]

            # Insertar los detalles de la factura
            for pid, nombre, cant, precio, item_subtotal in detalles:
                cur.execute("""
                    INSERT INTO detalle_factura (factura_id, producto_id, descripcion, cantidad, precio_unitario, subtotal)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (factura_id, pid, nombre, cant, precio, item_subtotal))

            conn.commit()
            flash('Factura creada exitosamente.', 'success')
            return redirect(url_for('ver_facturas'))

        except Exception as e:
            conn.rollback()
            flash(f'Error al crear factura: {e}', 'danger')

    else:
        # Método GET: Obtener el siguiente número de factura para autoincrementar
        cur.execute("SELECT MAX(CAST(numero_factura AS INTEGER)) FROM facturas")
        ultimo_num = cur.fetchone()[0]
        siguiente_numero = (ultimo_num or 0) + 1

    cur.close()
    conn.close()

    return render_template('facturas/nueva.html', contratos=contratos, productos=productos, siguiente_numero=siguiente_numero)

from flask import request

@app.route('/facturas/lista')
def ver_facturas():
    conn = get_db_connection()
    cur = conn.cursor()

    # Página actual, default 1
    page = request.args.get('page', 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page

    # Total facturas para calcular páginas
    cur.execute("SELECT COUNT(*) FROM facturas")
    total_facturas = cur.fetchone()[0]

    # Obtener facturas con paginación
    cur.execute("""
        SELECT f.id, f.numero_factura, f.fecha_emision, f.total, f.estado,
               c.nombres, c.apellidos, c.numero_contrato, f.comprobante_url
        FROM facturas f
        JOIN contratos c ON f.contrato_id = c.id
        ORDER BY f.fecha_emision DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    filas = cur.fetchall()

    facturas = []
    for fila in filas:
        facturas.append({
            'id': fila[0],
            'numero_factura': fila[1],
            'fecha_emision': fila[2],
            'total': float(fila[3]),
            'estado': fila[4],
            'cliente': f"{fila[5]} {fila[6]}",
            'numero_contrato': fila[7],
            'comprobante_url': fila[8]
        })

    cur.close()
    conn.close()

    total_pages = (total_facturas + per_page - 1) // per_page

    return render_template('facturas/lista.html', facturas=facturas, page=page, total_pages=total_pages)


@app.route('/facturas/<int:id>')
def ver_factura(id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener factura principal con datos del cliente
    cur.execute("""
        SELECT f.*, c.numero_contrato, c.nombres, c.apellidos
        FROM facturas f
        JOIN contratos c ON f.contrato_id = c.id
        WHERE f.id = %s
    """, (id,))
    factura = cur.fetchone()
    columnas_factura = [desc[0] for desc in cur.description]
    factura = dict(zip(columnas_factura, factura)) if factura else None

    # Obtener detalles de la factura
    cur.execute("""
        SELECT df.descripcion, df.cantidad, df.precio_unitario, df.subtotal
        FROM detalle_factura df
        WHERE df.factura_id = %s
    """, (id,))
    detalles = cur.fetchall()
    columnas_detalle = [desc[0] for desc in cur.description]
    detalles = [dict(zip(columnas_detalle, fila)) for fila in detalles]

    cur.close()
    conn.close()

    if not factura:
        flash("Factura no encontrada.", "danger")
        return redirect(url_for('ver_facturas'))

    return render_template('facturas/ver.html', factura=factura, detalles=detalles)

@app.route('/facturas/<int:id>/descargar_pdf')
def descargar_pdf_factura(id):
    # Crear carpeta si no existe
    os.makedirs('static/facturas', exist_ok=True)
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener datos de factura y cliente
    cur.execute("""
        SELECT f.numero_factura, f.fecha_emision, f.subtotal, f.impuestos, f.total, f.estado, f.notas,
               c.nombres, c.apellidos, c.numero_contrato, c.direccion, c.telefono, c.email
        FROM facturas f
        JOIN contratos c ON f.contrato_id = c.id
        WHERE f.id = %s
    """, (id,))
    factura = cur.fetchone()

    if not factura:
        cur.close()
        conn.close()
        return "Factura no encontrada", 404

    (numero_factura, fecha_emision, subtotal, impuestos, total, estado, notas,
     nombres, apellidos, numero_contrato, direccion, telefono, email) = factura

    # Obtener detalles de factura
    cur.execute("""
        SELECT descripcion, cantidad, precio_unitario, subtotal
        FROM detalle_factura
        WHERE factura_id = %s
    """, (id,))
    detalles = cur.fetchall()

    cur.close()
    conn.close()

    # Crear PDF
    pdf = FPDF()
    pdf.add_page()

    # Título
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, f"Factura #{numero_factura}", ln=True, align='C')

    # Datos cliente
    pdf.set_font("Arial", '', 12)
    pdf.cell(0, 8, f"Cliente: {nombres} {apellidos}", ln=True)
    pdf.cell(0, 8, f"Número Contrato: {numero_contrato}", ln=True)
    pdf.cell(0, 8, f"Dirección: {direccion}", ln=True)
    pdf.cell(0, 8, f"Teléfono: {telefono}", ln=True)
    pdf.cell(0, 8, f"Email: {email}", ln=True)
    pdf.cell(0, 8, f"Fecha: {fecha_emision.strftime('%Y-%m-%d')}", ln=True)
    pdf.cell(0, 8, f"Estado: {estado}", ln=True)

    pdf.ln(10)  # salto de línea

    # Tabla encabezados
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(80, 8, "Descripción", border=1)
    pdf.cell(30, 8, "Cantidad", border=1, align='C')
    pdf.cell(40, 8, "Precio Unitario", border=1, align='R')
    pdf.cell(40, 8, "Subtotal", border=1, align='R')
    pdf.ln()

    # Tabla contenido
    pdf.set_font("Arial", '', 12)
    for desc, cant, precio_unit, subtotal_item in detalles:
        pdf.cell(80, 8, desc, border=1)
        pdf.cell(30, 8, str(cant), border=1, align='C')
        pdf.cell(40, 8, f"${precio_unit:.2f}", border=1, align='R')
        pdf.cell(40, 8, f"${subtotal_item:.2f}", border=1, align='R')
        pdf.ln()

    pdf.ln(5)
    pdf.cell(0, 8, f"Subtotal: ${subtotal:.2f}", ln=True, align='R')
    pdf.cell(0, 8, f"Impuestos: ${impuestos:.2f}", ln=True, align='R')
    pdf.cell(0, 8, f"Total: ${total:.2f}", ln=True, align='R')

    if notas:
        pdf.ln(10)
        pdf.multi_cell(0, 8, f"Notas: {notas}")

    # Generar PDF en memoria (string binario)
    pdf_bytes = pdf.output(dest='S').encode('latin1')

    # Convertir a BytesIO para send_file
    pdf_output = io.BytesIO(pdf_bytes)
    pdf_output.seek(0)

    # Enviar PDF como descarga
    return send_file(
    pdf_output,
    download_name=f"Factura_{numero_factura}.pdf",
    as_attachment=True,
    mimetype='application/pdf'
)

@app.route('/factura/pagar/<int:id>', methods=['POST'])
def pagar_factura(id):
    file = request.files['comprobante_pago']
    if file and file.filename:
        filename = secure_filename(file.filename)
        ruta = os.path.join('comprobantes', filename)
        file.save(ruta)

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE facturas SET estado = %s, comprobante = %s WHERE id = %s", ('Pagado', filename, id))
        conn.commit()
        cur.close()
        conn.close()

    return redirect(url_for('ver_facturas'))

import os

ruta_comprobantes = os.path.join('static', 'comprobantes')
os.makedirs(ruta_comprobantes, exist_ok=True)

@app.route('/facturas/<int:id>/subir_comprobante', methods=['POST'])
def subir_comprobante(id):
    if 'comprobante' not in request.files:
        flash('No se ha seleccionado ningún archivo', 'danger')
        return redirect(url_for('ver_facturas'))

    file = request.files['comprobante']

    if file.filename == '':
        flash('No se ha seleccionado ningún archivo', 'danger')
        return redirect(url_for('ver_facturas'))

    # Validar extensión si quieres (pdf, jpg, png, etc)
    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    if not allowed_file(file.filename):
        flash('Tipo de archivo no permitido', 'danger')
        return redirect(url_for('ver_facturas'))

    # Carpeta para guardar comprobantes
    upload_folder = os.path.join('static', 'comprobantes')
    os.makedirs(upload_folder, exist_ok=True)

    filename = f"comprobante_factura_{id}_" + file.filename
    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    # Guardar nombre del archivo y cambiar estado a 'Pagado' en DB
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE facturas
        SET estado = 'Pagado', comprobante_url = %s
        WHERE id = %s
    """, (filename, id))
    conn.commit()
    cur.close()
    conn.close()

    flash('Comprobante subido correctamente y factura marcada como pagada.', 'success')
    return redirect(url_for('ver_facturas'))

if __name__ == '__main__':
    app.run(debug=True)
